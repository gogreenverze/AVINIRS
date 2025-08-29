from flask import Blueprint, jsonify, request, send_file
from datetime import datetime, timedelta
import uuid
import json
import os
import random
import tempfile
import io
from functools import wraps
from werkzeug.utils import secure_filename
from io import BytesIO

# Import image processing
try:
    from PIL import Image as PILImage
    IMAGE_SUPPORT = True
except ImportError:
    IMAGE_SUPPORT = False

# Import utilities
from utils import token_required, read_data, write_data, paginate_results, filter_data_by_tenant, check_tenant_access, transform_master_data, require_module_access

# Import Excel parsing library
try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Import extended master data functions
from .master_data_extended import (
    # Create functions
    create_patient_generic, create_profile_master_generic, create_method_master_generic,
    create_antibiotic_master_generic, create_organism_master_generic, create_unit_of_measurement_generic,
    create_specimen_master_generic, create_organism_vs_antibiotic_generic, create_container_master_generic,
    create_main_department_master_generic, create_department_settings_generic,
    create_authorization_settings_generic, create_print_order_generic,
    create_test_master_generic, create_sub_test_master_generic,
    # Update functions
    update_patient_generic, update_profile_master_generic, update_method_master_generic,
    update_antibiotic_master_generic, update_organism_master_generic, update_unit_of_measurement_generic,
    update_specimen_master_generic, update_organism_vs_antibiotic_generic, update_container_master_generic,
    update_main_department_master_generic, update_department_settings_generic,
    update_authorization_settings_generic, update_print_order_generic,
    update_test_master_generic, update_sub_test_master_generic,
    # Delete functions
    delete_test_master_generic, delete_sub_test_master_generic,
    delete_patient_generic, delete_profile_master_generic, delete_method_master_generic,
    delete_antibiotic_master_generic, delete_organism_master_generic, delete_unit_of_measurement_generic,
    delete_specimen_master_generic, delete_organism_vs_antibiotic_generic, delete_container_master_generic,
    delete_main_department_master_generic, delete_department_settings_generic,
    delete_authorization_settings_generic, delete_print_order_generic
)

admin_bp = Blueprint('admin', __name__)

def safe_float(value, default=0):
    """Safely convert value to float, handling strings and None"""
    try:
        if value is None or value == '':
            return default
        return float(value)
    except (ValueError, TypeError):
        return default

def calculate_dashboard_metrics(patients, samples, results, billings, inventory, invoices, user_role):
    """Calculate comprehensive dashboard metrics"""
    today = datetime.now().strftime('%Y-%m-%d')
    current_month = datetime.now().strftime('%Y-%m')
    last_7_days = [(datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(6, -1, -1)]

    # Patient Management Metrics
    total_patients = len(patients)
    today_patients = len([p for p in patients if p.get('created_at', '').startswith(today)])
    monthly_patients = len([p for p in patients if p.get('created_at', '').startswith(current_month)])

    # Sample & Test Metrics
    total_samples = len(samples)
    pending_samples = len([s for s in samples if s.get('status') == 'Pending'])
    completed_samples = len([s for s in samples if s.get('status') == 'Completed'])

    # Results Metrics
    total_results = len(results)
    pending_results = len([r for r in results if r.get('status') == 'Pending'])
    completed_results = len([r for r in results if r.get('status') == 'Completed'])

    # Financial Metrics - Use safe_float to handle string/numeric inconsistencies
    total_revenue = sum(safe_float(b.get('total_amount', 0)) for b in billings)
    monthly_revenue = sum(safe_float(b.get('total_amount', 0)) for b in billings if b.get('invoice_date', '').startswith(current_month))
    pending_payments = sum(safe_float(b.get('total_amount', 0)) for b in billings if b.get('payment_status') == 'Pending')

    # Invoice Metrics
    total_invoices = len(invoices)
    pending_invoices = len([i for i in invoices if i.get('status') == 'Pending'])
    paid_invoices = len([i for i in invoices if i.get('status') == 'Paid'])

    # Inventory Metrics
    total_inventory_items = len(inventory)
    low_stock_items = len([i for i in inventory if i.get('quantity', 0) <= i.get('reorder_level', 0)])
    out_of_stock_items = len([i for i in inventory if i.get('quantity', 0) == 0])

    # Daily trends for last 7 days
    daily_trends = []
    for date in last_7_days:
        day_patients = len([p for p in patients if p.get('created_at', '').startswith(date)])
        day_samples = len([s for s in samples if s.get('created_at', '').startswith(date)])
        day_revenue = sum(safe_float(b.get('total_amount', 0)) for b in billings if b.get('invoice_date', '').startswith(date))

        daily_trends.append({
            'date': date,
            'patients': day_patients,
            'samples': day_samples,
            'revenue': day_revenue
        })

    # Recent activities (last 10 items)
    recent_patients = sorted([p for p in patients], key=lambda x: x.get('created_at', ''), reverse=True)[:10]
    recent_samples = sorted([s for s in samples], key=lambda x: x.get('created_at', ''), reverse=True)[:10]
    recent_billings = sorted([b for b in billings], key=lambda x: x.get('created_at', ''), reverse=True)[:10]

    return {
        'overview': {
            'total_patients': total_patients,
            'today_patients': today_patients,
            'monthly_patients': monthly_patients,
            'total_samples': total_samples,
            'pending_samples': pending_samples,
            'completed_samples': completed_samples,
            'total_results': total_results,
            'pending_results': pending_results,
            'completed_results': completed_results,
            'total_revenue': total_revenue,
            'monthly_revenue': monthly_revenue,
            'pending_payments': pending_payments,
            'total_invoices': total_invoices,
            'pending_invoices': pending_invoices,
            'paid_invoices': paid_invoices,
            'total_inventory_items': total_inventory_items,
            'low_stock_items': low_stock_items,
            'out_of_stock_items': out_of_stock_items
        },
        'trends': {
            'daily_trends': daily_trends,
            'monthly_revenue': monthly_revenue,
            'revenue_growth': calculate_revenue_growth(billings, current_month)
        },
        'recent_activities': {
            'patients': recent_patients,
            'samples': recent_samples,
            'billings': recent_billings
        },
        'alerts': generate_dashboard_alerts(inventory, billings, samples, results),
        'ai_insights': generate_ai_insights(patients, samples, results, billings, inventory, user_role)
    }

def calculate_revenue_growth(billings, current_month):
    """Calculate revenue growth compared to previous month"""
    try:
        current_year, current_month_num = current_month.split('-')
        prev_month_num = int(current_month_num) - 1
        prev_year = current_year

        if prev_month_num == 0:
            prev_month_num = 12
            prev_year = str(int(current_year) - 1)

        prev_month = f"{prev_year}-{prev_month_num:02d}"

        current_revenue = sum(safe_float(b.get('total_amount', 0)) for b in billings if b.get('invoice_date', '').startswith(current_month))
        prev_revenue = sum(safe_float(b.get('total_amount', 0)) for b in billings if b.get('invoice_date', '').startswith(prev_month))

        if prev_revenue > 0:
            growth = ((current_revenue - prev_revenue) / prev_revenue) * 100
            return round(growth, 2)
        return 0
    except:
        return 0

def generate_dashboard_alerts(inventory, billings, samples, results):
    """Generate important alerts for dashboard"""
    alerts = []

    # Low stock alerts
    low_stock = [i for i in inventory if i.get('quantity', 0) <= i.get('reorder_level', 0)]
    if low_stock:
        alerts.append({
            'type': 'warning',
            'title': 'Low Stock Alert',
            'message': f'{len(low_stock)} items are running low on stock',
            'count': len(low_stock),
            'action': '/inventory?filter=low_stock'
        })

    # Pending results alerts
    pending_results = [r for r in results if r.get('status') == 'Pending']
    if len(pending_results) > 10:
        alerts.append({
            'type': 'info',
            'title': 'Pending Results',
            'message': f'{len(pending_results)} results are pending review',
            'count': len(pending_results),
            'action': '/results?status=pending'
        })

    # Overdue payments
    overdue_payments = [b for b in billings if b.get('payment_status') == 'Pending' and
                       b.get('due_date') and b.get('due_date') < datetime.now().strftime('%Y-%m-%d')]
    if overdue_payments:
        alerts.append({
            'type': 'danger',
            'title': 'Overdue Payments',
            'message': f'{len(overdue_payments)} payments are overdue',
            'count': len(overdue_payments),
            'action': '/billing?status=overdue'
        })

    return alerts

def generate_ai_insights(patients, samples, results, billings, inventory, user_role):
    """Generate AI-powered insights and recommendations"""
    insights = []

    # Patient volume insights
    today = datetime.now().strftime('%Y-%m-%d')
    today_patients = len([p for p in patients if p.get('created_at', '').startswith(today)])
    avg_daily_patients = len(patients) / max(30, 1)  # Average over last 30 days

    if today_patients > avg_daily_patients * 1.5:
        insights.append({
            'type': 'trend',
            'category': 'Patient Management',
            'title': 'High Patient Volume Detected',
            'description': f'Today\'s patient count ({today_patients}) is {int((today_patients/avg_daily_patients - 1) * 100)}% above average',
            'recommendation': 'Consider allocating additional staff or extending hours',
            'priority': 'high'
        })

    # Revenue insights
    current_month = datetime.now().strftime('%Y-%m')
    monthly_revenue = sum(safe_float(b.get('total_amount', 0)) for b in billings if b.get('invoice_date', '').startswith(current_month))

    if monthly_revenue > 0:
        insights.append({
            'type': 'financial',
            'category': 'Financial Performance',
            'title': 'Revenue Performance',
            'description': f'Current month revenue: ₹{monthly_revenue:,.2f}',
            'recommendation': 'Monitor daily revenue trends for optimization opportunities',
            'priority': 'medium'
        })

    # Inventory optimization
    low_stock_items = [i for i in inventory if i.get('quantity', 0) <= i.get('reorder_level', 0)]
    if low_stock_items:
        insights.append({
            'type': 'operational',
            'category': 'Inventory Management',
            'title': 'Inventory Optimization Needed',
            'description': f'{len(low_stock_items)} items need restocking',
            'recommendation': 'Review reorder levels and supplier lead times',
            'priority': 'high'
        })

    # Test efficiency insights
    pending_results_ratio = len([r for r in results if r.get('status') == 'Pending']) / max(len(results), 1)
    if pending_results_ratio > 0.3:
        insights.append({
            'type': 'operational',
            'category': 'Lab Efficiency',
            'title': 'Result Processing Bottleneck',
            'description': f'{int(pending_results_ratio * 100)}% of results are pending',
            'recommendation': 'Review lab workflow and consider additional resources',
            'priority': 'high'
        })

    return insights

# Enhanced Dashboard API
@admin_bp.route('/api/dashboard/comprehensive', methods=['GET'])
@token_required
def get_comprehensive_dashboard():
    """Get comprehensive dashboard data with role-based filtering"""
    try:
        user = request.current_user
        user_role = user.get('role')
        user_tenant_id = user.get('tenant_id')

        # Get all data files
        patients = read_data('patients.json')
        samples = read_data('samples.json')
        results = read_data('results.json')
        billings = read_data('billings.json')
        inventory = read_data('inventory.json')
        invoices = read_data('invoices.json')

        # Apply role-based filtering
        if user_role in ['admin', 'hub_admin']:
            # Admin and hub_admin see all data
            filtered_patients = patients
            filtered_samples = samples
            filtered_results = results
            filtered_billings = billings
            filtered_inventory = inventory
            filtered_invoices = invoices
        else:
            # Franchise admin and other roles see only their tenant data
            filtered_patients = filter_data_by_tenant(patients, user)
            filtered_samples = filter_data_by_tenant(samples, user)
            filtered_results = filter_data_by_tenant(results, user)
            filtered_billings = filter_data_by_tenant(billings, user)
            filtered_inventory = filter_data_by_tenant(inventory, user)
            filtered_invoices = filter_data_by_tenant(invoices, user)

        # Calculate dashboard metrics
        dashboard_data = calculate_dashboard_metrics(
            filtered_patients, filtered_samples, filtered_results,
            filtered_billings, filtered_inventory, filtered_invoices, user_role
        )

        return jsonify({
            'success': True,
            'data': dashboard_data,
            'user_context': {
                'role': user_role,
                'tenant_id': user_tenant_id,
                'access_level': 'system_wide' if user_role in ['admin', 'hub_admin'] else 'franchise_only'
            }
        }), 200

    except Exception as e:
        print(f"[DASHBOARD ERROR] {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error fetching dashboard data: {str(e)}'
        }), 500

# Admin Routes
@admin_bp.route('/api/admin/analytics', methods=['GET'])
@token_required
def get_analytics():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    # Get counts from various data files
    patients = read_data('patients.json')
    samples = read_data('samples.json')
    results = read_data('results.json')
    billings = read_data('billings.json')

    # Calculate monthly revenue
    current_month = datetime.now().strftime('%Y-%m')
    monthly_revenue = sum(
        safe_float(b.get('total_amount', 0))
        for b in billings
        if b.get('invoice_date', '').startswith(current_month)
    )

    # Get sample type distribution
    sample_types = read_data('sample_types.json')
    sample_type_counts = []

    for sample_type in sample_types:
        count = len([s for s in samples if s.get('sample_type_id') == sample_type.get('id')])
        sample_type_counts.append({
            'id': sample_type.get('id'),
            'type_name': sample_type.get('type_name'),
            'count': count
        })

    # Sort by count (descending)
    sample_type_counts = sorted(sample_type_counts, key=lambda x: x.get('count'), reverse=True)

    # Generate test statistics for the last 30 days
    test_stats = []
    for i in range(30):
        date = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        count = len([r for r in results if r.get('created_at', '').startswith(date)])
        test_stats.append({
            'date': date,
            'count': count
        })

    # Reverse to show oldest to newest
    test_stats.reverse()

    # Calculate test count
    test_count = len(results)

    return jsonify({
        'patient_count': len(patients),
        'sample_count': len(samples),
        'test_count': test_count,
        'monthly_revenue': monthly_revenue,
        'sample_types': sample_type_counts,
        'test_stats': test_stats
    })

@admin_bp.route('/api/admin/users', methods=['GET'])
@token_required
@require_module_access('USER_MANAGEMENT')
def get_users():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin', 'franchise_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    users = read_data('users.json')

    # Apply role-based filtering for users
    current_user_role = request.current_user.get('role')
    current_user_tenant_id = request.current_user.get('tenant_id')

    if current_user_role == 'franchise_admin':
        # Franchise admin can only see users from their own tenant
        users = [u for u in users if u.get('tenant_id') == current_user_tenant_id]
    elif current_user_role == 'hub_admin':
        # Hub admin can see users from all franchises and their own hub
        tenants = read_data('tenants.json')
        user_tenant = next((t for t in tenants if t.get('id') == current_user_tenant_id), None)
        if user_tenant and user_tenant.get('is_hub'):
            franchise_tenant_ids = [t.get('id') for t in tenants if not t.get('is_hub')]
            users = [u for u in users if u.get('tenant_id') in franchise_tenant_ids or u.get('tenant_id') == current_user_tenant_id]

    # Remove passwords from user objects
    for user in users:
        user.pop('password', None)

    # Add tenant information
    tenants = read_data('tenants.json')
    for user in users:
        tenant_id = user.get('tenant_id')
        if tenant_id:
            tenant = next((t for t in tenants if t.get('id') == tenant_id), None)
            if tenant:
                user['tenant'] = {
                    'id': tenant.get('id'),
                    'name': tenant.get('name'),
                    'site_code': tenant.get('site_code')
                }

    return jsonify(users)

@admin_bp.route('/api/admin/users/<int:id>', methods=['GET'])
@token_required
def get_user(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    users = read_data('users.json')
    user = next((u for u in users if u['id'] == id), None)

    if not user:
        return jsonify({'message': 'User not found'}), 404

    # Remove password from user object
    user_copy = user.copy()
    user_copy.pop('password', None)

    # Add tenant information
    tenant_id = user_copy.get('tenant_id')
    if tenant_id:
        tenants = read_data('tenants.json')
        tenant = next((t for t in tenants if t.get('id') == tenant_id), None)
        if tenant:
            user_copy['tenant'] = tenant

    return jsonify(user_copy)

@admin_bp.route('/api/admin/users', methods=['POST'])
@token_required
def create_user():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['username', 'password', 'email', 'first_name', 'last_name', 'role', 'tenant_id']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    users = read_data('users.json')

    # Check if username already exists
    if any(u.get('username') == data['username'] for u in users):
        return jsonify({'message': 'Username already exists'}), 400

    # Check if email already exists
    if any(u.get('email') == data['email'] for u in users):
        return jsonify({'message': 'Email already exists'}), 400

    # Generate new user ID
    new_id = 1
    if users:
        new_id = max(u['id'] for u in users) + 1

    # Create new user
    new_user = {
        'id': new_id,
        'username': data['username'],
        'password': data['password'],
        'email': data['email'],
        'first_name': data['first_name'],
        'last_name': data['last_name'],
        'role': data['role'],
        'tenant_id': data['tenant_id'],
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    users.append(new_user)
    write_data('users.json', users)

    # Remove password from response
    new_user_copy = new_user.copy()
    new_user_copy.pop('password', None)

    return jsonify(new_user_copy), 201

@admin_bp.route('/api/admin/users/<int:id>', methods=['PUT'])
@token_required
def update_user(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    users = read_data('users.json')
    user_index = next((i for i, u in enumerate(users) if u['id'] == id), None)

    if user_index is None:
        return jsonify({'message': 'User not found'}), 404

    # Update user fields
    user = users[user_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            user[key] = value

    user['updated_at'] = datetime.now().isoformat()

    # Save updated users
    write_data('users.json', users)

    # Remove password from response
    user_copy = user.copy()
    user_copy.pop('password', None)

    return jsonify(user_copy)

@admin_bp.route('/api/admin/users/<int:id>', methods=['DELETE'])
@token_required
def delete_user(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    # Prevent deleting self
    if id == request.current_user.get('id'):
        return jsonify({'message': 'Cannot delete your own account'}), 400

    users = read_data('users.json')
    user_index = next((i for i, u in enumerate(users) if u['id'] == id), None)

    if user_index is None:
        return jsonify({'message': 'User not found'}), 404

    # Delete user
    deleted_user = users.pop(user_index)
    write_data('users.json', users)

    return jsonify({'message': 'User deleted successfully'})

@admin_bp.route('/api/admin/franchises', methods=['GET'])
@token_required
def get_franchises():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    tenants = read_data('tenants.json')

    # For franchise management, return ALL tenants (including hub)
    # This allows proper management of all franchise locations
    return jsonify(tenants)

@admin_bp.route('/api/admin/franchises/<int:id>', methods=['GET'])
@token_required
def get_franchise(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    tenants = read_data('tenants.json')
    tenant = next((t for t in tenants if t['id'] == id), None)

    if not tenant:
        return jsonify({'message': 'Franchise not found'}), 404

    # If hub_admin, they can access their own hub and all franchises under it
    if request.current_user.get('role') == 'hub_admin':
        user_tenant_id = request.current_user.get('tenant_id')
        # Hub admin can access their own hub or any franchise (non-hub)
        if tenant.get('is_hub') and tenant.get('id') != user_tenant_id:
            return jsonify({'message': 'Unauthorized'}), 403

    return jsonify(tenant)

@admin_bp.route('/api/admin/settings', methods=['GET'])
@token_required
@require_module_access('SETTINGS')
def get_settings():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    # Default settings structure
    default_settings = {
        'general': {
            'site_name': 'AVINI Labs Management',
            'site_logo': '',
            'site_favicon': '',
            'contact_email': 'admin@rsavini.com',
            'contact_phone': '+91-9876543210',
            'address': '123 Lab Street, Medical District, City - 123456',
            'footer_text': '© 2024 AVINI Labs Management System',
            'timezone': 'Asia/Kolkata',
            'date_format': 'DD-MM-YYYY',
            'time_format': '12h'
        },
        'email': {
            'smtp_host': '',
            'smtp_port': '587',
            'smtp_username': '',
            'smtp_password': '',
            'smtp_encryption': 'tls',
            'from_email': 'noreply@rsavini.com',
            'from_name': 'AVINI Labs'
        },
        'security': {
            'password_min_length': 8,
            'password_expiry_days': 90,
            'max_login_attempts': 5,
            'lockout_time_minutes': 30,
            'session_timeout_minutes': 60,
            'enable_2fa': False
        },
        'lab': {
            'enable_sample_tracking': True,
            'enable_qc': True,
            'default_sample_validity_days': 7,
            'enable_auto_numbering': True,
            'sample_id_prefix': 'SAM',
            'result_id_prefix': 'RES'
        },
        'billing': {
            'enable_gst': True,
            'gst_number': '',
            'default_gst_rate': 18,
            'enable_discount': True,
            'max_discount_percent': 20,
            'invoice_prefix': 'INV',
            'payment_terms_days': 30
        }
    }

    try:
        # Try to read existing settings
        settings = read_data('settings.json')
        if not settings:
            settings = default_settings
    except:
        # If file doesn't exist, use default settings
        settings = default_settings

    return jsonify(settings)

@admin_bp.route('/api/admin/settings', methods=['PUT'])
@token_required
def update_settings():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    try:
        # Save settings to file
        write_data('settings.json', data)
        return jsonify({'message': 'Settings updated successfully'})
    except Exception as e:
        return jsonify({'message': f'Failed to update settings: {str(e)}'}), 500

# GST Configuration Master Routes
@admin_bp.route('/api/admin/gst-config', methods=['GET'])
@token_required
def get_gst_config():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin', 'franchise_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    gst_configs = read_data('gst_config.json')

    # Apply tenant-based filtering
    gst_configs = filter_data_by_tenant(gst_configs, request.current_user)

    return jsonify(gst_configs)

@admin_bp.route('/api/admin/gst-config', methods=['POST'])
@token_required
def create_gst_config():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['name', 'rate', 'applicable_from']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    gst_configs = read_data('gst_config.json')

    # Generate new ID
    new_id = 1
    if gst_configs:
        new_id = max(config['id'] for config in gst_configs) + 1

    # Create new GST configuration
    new_config = {
        'id': new_id,
        'name': data['name'],
        'description': data.get('description', ''),
        'rate': float(data['rate']),
        'applicable_from': data['applicable_from'],
        'applicable_to': data.get('applicable_to', ''),
        'is_default': data.get('is_default', False),
        'is_active': data.get('is_active', True),
        'tenant_id': request.current_user.get('tenant_id'),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    # If this is set as default, unset other defaults for this tenant
    if new_config['is_default']:
        for config in gst_configs:
            if config.get('tenant_id') == new_config['tenant_id']:
                config['is_default'] = False

    gst_configs.append(new_config)
    write_data('gst_config.json', gst_configs)

    return jsonify(new_config), 201

@admin_bp.route('/api/admin/gst-config/<int:config_id>', methods=['PUT'])
@token_required
def update_gst_config(config_id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()
    gst_configs = read_data('gst_config.json')

    # Find the configuration
    config = next((c for c in gst_configs if c['id'] == config_id), None)
    if not config:
        return jsonify({'message': 'GST configuration not found'}), 404

    # Check tenant access
    if not check_tenant_access(request.current_user, config.get('tenant_id')):
        return jsonify({'message': 'Unauthorized access to this GST configuration'}), 403

    # Update fields
    config.update({
        'name': data.get('name', config['name']),
        'description': data.get('description', config['description']),
        'rate': float(data.get('rate', config['rate'])),
        'applicable_from': data.get('applicable_from', config['applicable_from']),
        'applicable_to': data.get('applicable_to', config['applicable_to']),
        'is_default': data.get('is_default', config['is_default']),
        'is_active': data.get('is_active', config['is_active']),
        'updated_at': datetime.now().isoformat()
    })

    # If this is set as default, unset other defaults for this tenant
    if config['is_default']:
        for other_config in gst_configs:
            if other_config['id'] != config_id and other_config.get('tenant_id') == config['tenant_id']:
                other_config['is_default'] = False

    write_data('gst_config.json', gst_configs)
    return jsonify(config)

@admin_bp.route('/api/admin/gst-config/<int:config_id>', methods=['DELETE'])
@token_required
def delete_gst_config(config_id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    gst_configs = read_data('gst_config.json')

    # Find the configuration
    config = next((c for c in gst_configs if c['id'] == config_id), None)
    if not config:
        return jsonify({'message': 'GST configuration not found'}), 404

    # Check tenant access
    if not check_tenant_access(request.current_user, config.get('tenant_id')):
        return jsonify({'message': 'Unauthorized access to this GST configuration'}), 403

    # Remove the configuration
    gst_configs = [c for c in gst_configs if c['id'] != config_id]
    write_data('gst_config.json', gst_configs)

    return jsonify({'message': 'GST configuration deleted successfully'})

# Master Data API
@admin_bp.route('/api/admin/master-data', methods=['GET'])
@token_required
def get_master_data():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        # Get counts and basic info for all master data entities
        doctors = read_data('doctors.json')
        test_categories = read_data('test_categories.json')
        test_parameters = read_data('test_parameters.json')
        departments = read_data('departments.json')
        payment_methods = read_data('payment_methods.json')
        containers = read_data('containers.json')
        instruments = read_data('instruments.json')
        reagents = read_data('reagents.json')
        suppliers = read_data('suppliers.json')
        units = read_data('units.json')
        test_methods = read_data('test_methods.json')
        tests = read_data('tests.json')
        test_panels = read_data('test_panels.json')
        sample_types = read_data('sample_types.json')

        # New master data categories from Excel
        patients = read_data('patients.json')
        profile_master = read_data('profile_master.json')
        method_master = read_data('method_master.json')
        antibiotic_master = read_data('antibiotic_master.json')
        organism_master = read_data('organism_master.json')
        unit_of_measurement = read_data('unit_of_measurement.json')
        specimen_master = read_data('specimen_master.json')
        organism_vs_antibiotic = read_data('organism_vs_antibiotic.json')
        container_master = read_data('container_master.json')
        main_department_master = read_data('main_department_master.json')
        department_settings = read_data('department_settings.json')
        authorization_settings = read_data('authorization_settings.json')
        print_order = read_data('print_order.json')
        test_master = read_data('test_master.json')
        sub_test_master = read_data('sub_test_master.json')
        profile_data = read_data('profile_data.json')

        # Apply data transformations to fix field name mismatches
        master_data = {
            # Original categories - apply transformations
            'doctors': transform_master_data(doctors, 'doctors'),
            'testCategories': transform_master_data(test_categories, 'testCategories'),
            'testParameters': transform_master_data(test_parameters, 'testParameters'),
            'sampleTypes': transform_master_data(sample_types, 'sampleTypes'),
            'departments': transform_master_data(departments, 'departments'),
            'paymentMethods': transform_master_data(payment_methods, 'paymentMethods'),
            'containers': transform_master_data(containers, 'containers'),
            'instruments': transform_master_data(instruments, 'instruments'),
            'reagents': transform_master_data(reagents, 'reagents'),
            'suppliers': transform_master_data(suppliers, 'suppliers'),
            'units': transform_master_data(units, 'units'),
            'testMethods': transform_master_data(test_methods, 'testMethods'),
            'tests': transform_master_data(tests, 'tests'),
            'test_panels': transform_master_data(test_panels, 'test_panels'),
            # New categories from Excel - apply transformations
            'patients': transform_master_data(patients, 'patients'),
            'profileMaster': transform_master_data(profile_master, 'profileMaster'),
            'methodMaster': transform_master_data(method_master, 'methodMaster'),
            'antibioticMaster': transform_master_data(antibiotic_master, 'antibioticMaster'),
            'organismMaster': transform_master_data(organism_master, 'organismMaster'),
            'unitOfMeasurement': transform_master_data(unit_of_measurement, 'unitOfMeasurement'),
            'specimenMaster': transform_master_data(specimen_master, 'specimenMaster'),
            'organismVsAntibiotic': transform_master_data(organism_vs_antibiotic, 'organismVsAntibiotic'),
            'containerMaster': transform_master_data(container_master, 'containerMaster'),
            'mainDepartmentMaster': transform_master_data(main_department_master, 'mainDepartmentMaster'),
            'departmentSettings': transform_master_data(department_settings, 'departmentSettings'),
            'authorizationSettings': transform_master_data(authorization_settings, 'authorizationSettings'),
            'printOrder': transform_master_data(print_order, 'printOrder'),
            'testMaster': transform_master_data(test_master, 'testMaster'),
            'subTestMaster': transform_master_data(sub_test_master, 'subTestMaster'),
            'profileData': transform_master_data(profile_data, 'profileData')
        }

        return jsonify(master_data)

    except Exception as e:
        return jsonify({'message': f'Failed to load master data: {str(e)}'}), 500

# Excel Reference Data API
@admin_bp.route('/api/admin/excel-reference-data', methods=['GET'])
@token_required
def get_excel_reference_data():
    """
    Parse Excel file server-side and return JSON data with normalized test name keys
    for efficient lookup in the frontend Result Master form.
    """
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    if not EXCEL_SUPPORT:
        # Return sample data structure for testing when openpyxl is not available
        print("Warning: openpyxl not installed. Returning sample Excel reference data for testing.")
        sample_data = {
            "gad-65 - igg": {
                "testName": "GAD-65 - IgG",
                "code": "123",
                "department": "IMMUNOLOGY",
                "referenceRange": "< 1.0 U/ml",
                "resultUnit": "U/ml",
                "decimals": "1",
                "notes": "Autoimmune marker for diabetes",
                "criticalLow": "",
                "criticalHigh": "",
                "sheetName": "IMMUNOLOGY"
            },
            "25 hydroxy vitamin d3": {
                "testName": "25 Hydroxy Vitamin D3",
                "code": "630",
                "department": "IMMUNOLOGY",
                "referenceRange": "< 20 : Deficiency, 20-29 : Insufficiency, 30-100 : Sufficiency, > 100 : Toxicity",
                "resultUnit": "ng/ml",
                "decimals": "1",
                "notes": "Vitamin D status assessment",
                "criticalLow": "",
                "criticalHigh": "",
                "sheetName": "IMMUNOLOGY"
            },
            "1,25 dihydroxyvitamin d": {
                "testName": "1,25 Dihydroxyvitamin D",
                "code": "648",
                "department": "IMMUNOLOGY",
                "referenceRange": "19.9 - 79.3 pg/ml",
                "resultUnit": "pg/ml",
                "decimals": "1",
                "notes": "Active form of vitamin D",
                "criticalLow": "",
                "criticalHigh": "",
                "sheetName": "IMMUNOLOGY"
            }
        }
        return jsonify(sample_data)

    try:
        # Path to the Excel file
        excel_file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '..', 'public', 'dynamic data fetch.xlsx')

        if not os.path.exists(excel_file_path):
            return jsonify({'message': 'Excel file not found'}), 404

        # Load the workbook
        workbook = load_workbook(excel_file_path, data_only=True)

        reference_data = {}

        # Helper function to extract value from row with multiple possible column names
        def get_value_from_row(row_dict, possible_keys):
            for key in possible_keys:
                if key in row_dict and row_dict[key] is not None and str(row_dict[key]).strip():
                    return str(row_dict[key]).strip()
            return ''

        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get header row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append('')

            # Process data rows
            for row_num in range(2, sheet.max_row + 1):
                row_data = {}

                # Create row dictionary
                for col_num, header in enumerate(headers, 1):
                    if header:
                        cell_value = sheet.cell(row=row_num, column=col_num).value
                        row_data[header] = cell_value if cell_value is not None else ''

                # Extract test name with multiple possible column names
                test_name = get_value_from_row(row_data, [
                    'Test Name',
                    'Test Name ',
                    'TestName',
                    'test_name',
                    'Test'
                ])

                if test_name:
                    # Normalize the key for consistent lookup
                    normalized_key = test_name.lower().strip()

                    reference_data[normalized_key] = {
                        'testName': test_name,
                        # Enhanced code extraction
                        'code': get_value_from_row(row_data, [
                            'code',
                            'Code',
                            'TEST_CODE',
                            'Test Code',
                            'TestCode'
                        ]),
                        # Enhanced department extraction
                        'department': get_value_from_row(row_data, [
                            'Department',
                            'department',
                            'DEPARTMENT',
                            'Dept'
                        ]) or sheet_name,
                        # Enhanced reference range extraction
                        'referenceRange': get_value_from_row(row_data, [
                            'Referance Range',
                            'Reference Range',
                            '\r\nReference Range',
                            'ReferenceRange',
                            'Normal Range',
                            'Range'
                        ]),
                        # Enhanced result unit extraction
                        'resultUnit': get_value_from_row(row_data, [
                            'Result Unit',
                            'Result unit',
                            'ResultUnit',
                            'Unit',
                            'Units'
                        ]),
                        # Enhanced decimals extraction
                        'decimals': get_value_from_row(row_data, [
                            'No of decimals',
                            'No. of Decimals',
                            '\tNo. of Decimals',
                            'Decimals',
                            'Decimal Places',
                            'DecimalPlaces'
                        ]),
                        # Enhanced notes extraction
                        'notes': get_value_from_row(row_data, [
                            'Notes',
                            'Note',
                            'Comments',
                            'Comment',
                            'Remarks'
                        ]),
                        # Enhanced critical values extraction
                        'criticalLow': get_value_from_row(row_data, [
                            'Critical Low',
                            'CriticalLow',
                            'Low Critical',
                            'Panic Low'
                        ]),
                        'criticalHigh': get_value_from_row(row_data, [
                            'Critical High',
                            'CriticalHigh',
                            'High Critical',
                            'Panic High'
                        ]),
                        'sheetName': sheet_name
                    }

        # Log the number of entries processed
        print(f"Excel reference data loaded: {len(reference_data)} entries from {len(workbook.sheetnames)} sheets")

        return jsonify(reference_data)

    except Exception as e:
        print(f"Error parsing Excel file: {str(e)}")
        return jsonify({'message': f'Failed to parse Excel file: {str(e)}'}), 500

# Doctor Management Routes
@admin_bp.route('/api/admin/doctors', methods=['GET'])
@token_required
def get_doctors():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    doctors = read_data('doctors.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    doctors = sorted(doctors, key=lambda x: x.get('created_at', ''), reverse=True)

    # Paginate results
    paginated_data = paginate_results(doctors, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/doctors/<int:id>', methods=['GET'])
@token_required
def get_doctor(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    doctors = read_data('doctors.json')
    doctor = next((d for d in doctors if d['id'] == id), None)

    if not doctor:
        return jsonify({'message': 'Doctor not found'}), 404

    return jsonify(doctor)

@admin_bp.route('/api/admin/doctors', methods=['POST'])
@token_required
def create_doctor():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['first_name', 'last_name', 'phone']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    doctors = read_data('doctors.json')

    # Generate new doctor ID
    new_id = 1
    if doctors:
        new_id = max(d['id'] for d in doctors) + 1

    # Create new doctor
    new_doctor = {
        'id': new_id,
        'first_name': data['first_name'],
        'last_name': data['last_name'],
        'email': data.get('email', ''),
        'phone': data['phone'],
        'specialty': data.get('specialty', ''),
        'qualification': data.get('qualification', ''),
        'license_number': data.get('license_number', ''),
        'address': data.get('address', ''),
        'city': data.get('city', ''),
        'state': data.get('state', ''),
        'pincode': data.get('pincode', ''),
        'status': data.get('status', 'Active'),
        'consultation_fee': data.get('consultation_fee', 0),
        'experience_years': data.get('experience_years', 0),
        'notes': data.get('notes', ''),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    doctors.append(new_doctor)
    write_data('doctors.json', doctors)

    return jsonify(new_doctor), 201

@admin_bp.route('/api/admin/doctors/<int:id>', methods=['PUT'])
@token_required
def update_doctor(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    doctors = read_data('doctors.json')
    doctor_index = next((i for i, d in enumerate(doctors) if d['id'] == id), None)

    if doctor_index is None:
        return jsonify({'message': 'Doctor not found'}), 404

    # Update doctor fields
    doctor = doctors[doctor_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            doctor[key] = value

    doctor['updated_at'] = datetime.now().isoformat()

    # Save updated doctors
    write_data('doctors.json', doctors)

    return jsonify(doctor)

@admin_bp.route('/api/admin/doctors/<int:id>', methods=['DELETE'])
@token_required
def delete_doctor(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    doctors = read_data('doctors.json')
    doctor_index = next((i for i, d in enumerate(doctors) if d['id'] == id), None)

    if doctor_index is None:
        return jsonify({'message': 'Doctor not found'}), 404

    # Delete doctor
    deleted_doctor = doctors.pop(doctor_index)
    write_data('doctors.json', doctors)

    return jsonify({'message': 'Doctor deleted successfully'})

# Test Category Management Routes
@admin_bp.route('/api/admin/test-categories', methods=['GET'])
@token_required
def get_test_categories():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    categories = read_data('test_categories.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    categories = sorted(categories, key=lambda x: x.get('created_at', ''), reverse=True)

    # Paginate results
    paginated_data = paginate_results(categories, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/test-categories/<int:id>', methods=['GET'])
@token_required
def get_test_category(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    categories = read_data('test_categories.json')
    category = next((c for c in categories if c['id'] == id), None)

    if not category:
        return jsonify({'message': 'Test category not found'}), 404

    return jsonify(category)

@admin_bp.route('/api/admin/test-categories', methods=['POST'])
@token_required
def create_test_category():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['name', 'code']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    categories = read_data('test_categories.json')

    # Check if code already exists
    if any(c.get('code') == data['code'].upper() for c in categories):
        return jsonify({'message': 'Category code already exists'}), 400

    # Generate new category ID
    new_id = 1
    if categories:
        new_id = max(c['id'] for c in categories) + 1

    # Create new category
    new_category = {
        'id': new_id,
        'name': data['name'],
        'code': data['code'].upper(),
        'description': data.get('description', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    categories.append(new_category)
    write_data('test_categories.json', categories)

    return jsonify(new_category), 201

@admin_bp.route('/api/admin/test-categories/<int:id>', methods=['PUT'])
@token_required
def update_test_category(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    categories = read_data('test_categories.json')
    category_index = next((i for i, c in enumerate(categories) if c['id'] == id), None)

    if category_index is None:
        return jsonify({'message': 'Test category not found'}), 404

    # Update category fields
    category = categories[category_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'code':
                category[key] = value.upper()
            else:
                category[key] = value

    category['updated_at'] = datetime.now().isoformat()

    # Save updated categories
    write_data('test_categories.json', categories)

    return jsonify(category)

@admin_bp.route('/api/admin/test-categories/<int:id>', methods=['DELETE'])
@token_required
def delete_test_category(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    categories = read_data('test_categories.json')
    category_index = next((i for i, c in enumerate(categories) if c['id'] == id), None)

    if category_index is None:
        return jsonify({'message': 'Test category not found'}), 404

    # Delete category
    deleted_category = categories.pop(category_index)
    write_data('test_categories.json', categories)

    return jsonify({'message': 'Test category deleted successfully'})

# Test Management Routes
@admin_bp.route('/api/admin/tests', methods=['GET'])
@token_required
def get_tests():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    tests = read_data('tests.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    tests = sorted(tests, key=lambda x: x.get('created_at', ''), reverse=True)

    # Add category information
    categories = read_data('test_categories.json')
    for test in tests:
        category_id = test.get('category_id')
        if category_id:
            category = next((c for c in categories if c.get('id') == category_id), None)
            if category:
                test['category'] = {
                    'id': category.get('id'),
                    'name': category.get('name'),
                    'code': category.get('code')
                }

    # Paginate results
    paginated_data = paginate_results(tests, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/tests/<int:id>', methods=['GET'])
@token_required
def get_test(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    tests = read_data('tests.json')
    test = next((t for t in tests if t['id'] == id), None)

    if not test:
        return jsonify({'message': 'Test not found'}), 404

    # Add category information
    category_id = test.get('category_id')
    if category_id:
        categories = read_data('test_categories.json')
        category = next((c for c in categories if c.get('id') == category_id), None)
        if category:
            test['category'] = category

    return jsonify(test)

@admin_bp.route('/api/admin/tests', methods=['POST'])
@token_required
def create_test():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['test_name', 'test_code', 'category_id']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    tests = read_data('tests.json')

    # Check if test code already exists
    if any(t.get('test_code') == data['test_code'].upper() for t in tests):
        return jsonify({'message': 'Test code already exists'}), 400

    # Generate new test ID
    new_id = 1
    if tests:
        new_id = max(t['id'] for t in tests) + 1

    # Create new test
    new_test = {
        'id': new_id,
        'test_name': data['test_name'],
        'test_code': data['test_code'].upper(),
        'category_id': data['category_id'],
        'price': data.get('price', 0),
        'normal_range': data.get('normal_range', ''),
        'unit': data.get('unit', ''),
        'method': data.get('method', ''),
        'sample_type': data.get('sample_type', ''),
        'turnaround_time': data.get('turnaround_time', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    tests.append(new_test)
    write_data('tests.json', tests)

    return jsonify(new_test), 201

@admin_bp.route('/api/admin/tests/<int:id>', methods=['PUT'])
@token_required
def update_test(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    tests = read_data('tests.json')
    test_index = next((i for i, t in enumerate(tests) if t['id'] == id), None)

    if test_index is None:
        return jsonify({'message': 'Test not found'}), 404

    # Update test fields
    test = tests[test_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'test_code':
                test[key] = value.upper()
            else:
                test[key] = value

    test['updated_at'] = datetime.now().isoformat()

    # Save updated tests
    write_data('tests.json', tests)

    return jsonify(test)

@admin_bp.route('/api/admin/tests/<int:id>', methods=['DELETE'])
@token_required
def delete_test(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    tests = read_data('tests.json')
    test_index = next((i for i, t in enumerate(tests) if t['id'] == id), None)

    if test_index is None:
        return jsonify({'message': 'Test not found'}), 404

    # Delete test
    deleted_test = tests.pop(test_index)
    write_data('tests.json', tests)

    return jsonify({'message': 'Test deleted successfully'})

# Test Panel Management Routes
@admin_bp.route('/api/admin/test-panels', methods=['GET'])
@token_required
def get_test_panels():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    panels = read_data('test_panels.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    panels = sorted(panels, key=lambda x: x.get('created_at', ''), reverse=True)

    # Add test information
    tests = read_data('tests.json')
    for panel in panels:
        test_ids = panel.get('test_ids', [])
        if test_ids:
            panel_tests = [t for t in tests if t.get('id') in test_ids]
            panel['tests'] = panel_tests

    # Paginate results
    paginated_data = paginate_results(panels, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/test-panels', methods=['POST'])
@token_required
def create_test_panel():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['name', 'code']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    panels = read_data('test_panels.json')

    # Check if code already exists
    if any(p.get('code') == data['code'].upper() for p in panels):
        return jsonify({'message': 'Panel code already exists'}), 400

    # Generate new panel ID
    new_id = 1
    if panels:
        new_id = max(p['id'] for p in panels) + 1

    # Create new panel
    new_panel = {
        'id': new_id,
        'name': data['name'],
        'code': data['code'].upper(),
        'description': data.get('description', ''),
        'price': data.get('price', 0),
        'test_ids': data.get('test_ids', []),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    panels.append(new_panel)
    write_data('test_panels.json', panels)

    return jsonify(new_panel), 201

@admin_bp.route('/api/admin/test-panels/<int:id>', methods=['PUT'])
@token_required
def update_test_panel(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    panels = read_data('test_panels.json')
    panel_index = next((i for i, p in enumerate(panels) if p['id'] == id), None)

    if panel_index is None:
        return jsonify({'message': 'Test panel not found'}), 404

    # Update panel fields
    panel = panels[panel_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'code':
                panel[key] = value.upper()
            else:
                panel[key] = value

    panel['updated_at'] = datetime.now().isoformat()

    # Save updated panels
    write_data('test_panels.json', panels)

    return jsonify(panel)

@admin_bp.route('/api/admin/test-panels/<int:id>', methods=['DELETE'])
@token_required
def delete_test_panel(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    panels = read_data('test_panels.json')
    panel_index = next((i for i, p in enumerate(panels) if p['id'] == id), None)

    if panel_index is None:
        return jsonify({'message': 'Test panel not found'}), 404

    # Delete panel
    deleted_panel = panels.pop(panel_index)
    write_data('test_panels.json', panels)

    return jsonify({'message': 'Test panel deleted successfully'})

# Container Management Routes
@admin_bp.route('/api/admin/containers', methods=['GET'])
@token_required
def get_containers():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    containers = read_data('containers.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    containers = sorted(containers, key=lambda x: x.get('created_at', ''), reverse=True)

    # Paginate results
    paginated_data = paginate_results(containers, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/containers', methods=['POST'])
@token_required
def create_container():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['name', 'code']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    containers = read_data('containers.json')

    # Check if code already exists
    if any(c.get('code') == data['code'].upper() for c in containers):
        return jsonify({'message': 'Container code already exists'}), 400

    # Generate new container ID
    new_id = 1
    if containers:
        new_id = max(c['id'] for c in containers) + 1

    # Create new container
    new_container = {
        'id': new_id,
        'name': data['name'],
        'code': data['code'].upper(),
        'description': data.get('description', ''),
        'volume': data.get('volume', ''),
        'color': data.get('color', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    containers.append(new_container)
    write_data('containers.json', containers)

    return jsonify(new_container), 201

@admin_bp.route('/api/admin/containers/<int:id>', methods=['PUT'])
@token_required
def update_container(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    containers = read_data('containers.json')
    container_index = next((i for i, c in enumerate(containers) if c['id'] == id), None)

    if container_index is None:
        return jsonify({'message': 'Container not found'}), 404

    # Update container fields
    container = containers[container_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'code':
                container[key] = value.upper()
            else:
                container[key] = value

    container['updated_at'] = datetime.now().isoformat()

    # Save updated containers
    write_data('containers.json', containers)

    return jsonify(container)

@admin_bp.route('/api/admin/containers/<int:id>', methods=['DELETE'])
@token_required
def delete_container(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    containers = read_data('containers.json')
    container_index = next((i for i, c in enumerate(containers) if c['id'] == id), None)

    if container_index is None:
        return jsonify({'message': 'Container not found'}), 404

    # Delete container
    deleted_container = containers.pop(container_index)
    write_data('containers.json', containers)

    return jsonify({'message': 'Container deleted successfully'})

# Role Management Routes
@admin_bp.route('/api/admin/roles', methods=['GET'])
@token_required
def get_roles():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    roles = read_data('roles.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    roles = sorted(roles, key=lambda x: x.get('created_at', ''), reverse=True)

    # Add permission information
    permissions = read_data('permissions.json')
    for role in roles:
        permission_ids = role.get('permission_ids', [])
        if permission_ids:
            role_permissions = [p for p in permissions if p.get('id') in permission_ids]
            role['permissions'] = role_permissions

    # Paginate results
    paginated_data = paginate_results(roles, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/roles', methods=['POST'])
@token_required
def create_role():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['name', 'code']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    roles = read_data('roles.json')

    # Check if code already exists
    if any(r.get('code') == data['code'].upper() for r in roles):
        return jsonify({'message': 'Role code already exists'}), 400

    # Generate new role ID
    new_id = 1
    if roles:
        new_id = max(r['id'] for r in roles) + 1

    # Create new role
    new_role = {
        'id': new_id,
        'name': data['name'],
        'code': data['code'].upper(),
        'description': data.get('description', ''),
        'permission_ids': data.get('permission_ids', []),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    roles.append(new_role)
    write_data('roles.json', roles)

    return jsonify(new_role), 201

@admin_bp.route('/api/admin/roles/<int:id>', methods=['PUT'])
@token_required
def update_role(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    roles = read_data('roles.json')
    role_index = next((i for i, r in enumerate(roles) if r['id'] == id), None)

    if role_index is None:
        return jsonify({'message': 'Role not found'}), 404

    # Update role fields
    role = roles[role_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'code':
                role[key] = value.upper()
            else:
                role[key] = value

    role['updated_at'] = datetime.now().isoformat()

    # Save updated roles
    write_data('roles.json', roles)

    return jsonify(role)

@admin_bp.route('/api/admin/roles/<int:id>', methods=['DELETE'])
@token_required
def delete_role(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    roles = read_data('roles.json')
    role_index = next((i for i, r in enumerate(roles) if r['id'] == id), None)

    if role_index is None:
        return jsonify({'message': 'Role not found'}), 404

    # Delete role
    deleted_role = roles.pop(role_index)
    write_data('roles.json', roles)

    return jsonify({'message': 'Role deleted successfully'})

# Permission Management Routes
@admin_bp.route('/api/admin/permissions', methods=['GET'])
@token_required
def get_permissions():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    permissions = read_data('permissions.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    permissions = sorted(permissions, key=lambda x: x.get('created_at', ''), reverse=True)

    # Paginate results
    paginated_data = paginate_results(permissions, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/permissions', methods=['POST'])
@token_required
def create_permission():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['name', 'code']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    permissions = read_data('permissions.json')

    # Check if code already exists
    if any(p.get('code') == data['code'].upper() for p in permissions):
        return jsonify({'message': 'Permission code already exists'}), 400

    # Generate new permission ID
    new_id = 1
    if permissions:
        new_id = max(p['id'] for p in permissions) + 1

    # Create new permission
    new_permission = {
        'id': new_id,
        'name': data['name'],
        'code': data['code'].upper(),
        'description': data.get('description', ''),
        'module': data.get('module', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    permissions.append(new_permission)
    write_data('permissions.json', permissions)

    return jsonify(new_permission), 201

@admin_bp.route('/api/admin/permissions/<int:id>', methods=['PUT'])
@token_required
def update_permission(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    permissions = read_data('permissions.json')
    permission_index = next((i for i, p in enumerate(permissions) if p['id'] == id), None)

    if permission_index is None:
        return jsonify({'message': 'Permission not found'}), 404

    # Update permission fields
    permission = permissions[permission_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'code':
                permission[key] = value.upper()
            else:
                permission[key] = value

    permission['updated_at'] = datetime.now().isoformat()

    # Save updated permissions
    write_data('permissions.json', permissions)

    return jsonify(permission)

@admin_bp.route('/api/admin/permissions/<int:id>', methods=['DELETE'])
@token_required
def delete_permission(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    permissions = read_data('permissions.json')
    permission_index = next((i for i, p in enumerate(permissions) if p['id'] == id), None)

    if permission_index is None:
        return jsonify({'message': 'Permission not found'}), 404

    # Delete permission
    deleted_permission = permissions.pop(permission_index)
    write_data('permissions.json', permissions)

    return jsonify({'message': 'Permission deleted successfully'})

# Franchise Management Routes
@admin_bp.route('/api/admin/franchises', methods=['POST'])
@token_required
def create_franchise():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['name', 'site_code', 'contact_phone']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    tenants = read_data('tenants.json')

    # Check if site code already exists
    if any(t.get('site_code') == data['site_code'].upper() for t in tenants):
        return jsonify({'message': 'Site code already exists'}), 400

    # Generate new franchise ID
    new_id = 1
    if tenants:
        new_id = max(t['id'] for t in tenants) + 1

    # Create new franchise
    new_franchise = {
        'id': new_id,
        'name': data['name'],
        'site_code': data['site_code'].upper(),
        'address': data.get('address', ''),
        'city': data.get('city', ''),
        'state': data.get('state', 'Tamil Nadu'),
        'pincode': data.get('pincode', ''),
        'contact_phone': data['contact_phone'],
        'email': data.get('email', ''),
        'license_number': data.get('license_number', ''),
        'established_date': data.get('established_date', ''),
        'is_hub': data.get('is_hub', False),
        'is_active': data.get('is_active', True),
        'use_site_code_prefix': data.get('use_site_code_prefix', True),  # Default to True for backward compatibility
        'franchise_fee': data.get('franchise_fee', 0),
        'monthly_fee': data.get('monthly_fee', 0),
        'commission_rate': data.get('commission_rate', 0),
        'contact_person': data.get('contact_person', ''),
        'contact_person_phone': data.get('contact_person_phone', ''),
        'notes': data.get('notes', ''),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    tenants.append(new_franchise)
    write_data('tenants.json', tenants)

    return jsonify(new_franchise), 201

@admin_bp.route('/api/admin/franchises/<int:id>', methods=['PUT'])
@token_required
def update_franchise(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()
    tenants = read_data('tenants.json')

    # Find the franchise
    tenant_index = next((i for i, t in enumerate(tenants) if t['id'] == id), None)
    if tenant_index is None:
        return jsonify({'message': 'Franchise not found'}), 404

    # Check if site code already exists (excluding current franchise)
    if 'site_code' in data:
        if any(t.get('site_code') == data['site_code'].upper() and t['id'] != id for t in tenants):
            return jsonify({'message': 'Site code already exists'}), 400

    # Update franchise fields
    franchise = tenants[tenant_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'site_code':
                franchise[key] = value.upper()
            else:
                franchise[key] = value

    franchise['updated_at'] = datetime.now().isoformat()

    # Save updated tenants
    write_data('tenants.json', tenants)
    return jsonify(franchise)

@admin_bp.route('/api/admin/franchises/<int:id>', methods=['DELETE'])
@token_required
def delete_franchise(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    tenants = read_data('tenants.json')

    # Find the franchise
    tenant_index = next((i for i, t in enumerate(tenants) if t['id'] == id), None)
    if tenant_index is None:
        return jsonify({'message': 'Franchise not found'}), 404

    # Check if it's the hub - prevent deletion of hub
    franchise = tenants[tenant_index]
    if franchise.get('is_hub'):
        return jsonify({'message': 'Cannot delete hub franchise'}), 400

    # Delete franchise
    deleted_franchise = tenants.pop(tenant_index)
    write_data('tenants.json', tenants)

    return jsonify({'message': 'Franchise deleted successfully'})

# Sample Type Management Routes
@admin_bp.route('/api/admin/sample-types', methods=['GET'])
@token_required
def get_sample_types():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    sample_types = read_data('sample_types.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    sample_types = sorted(sample_types, key=lambda x: x.get('created_at', ''), reverse=True)

    # Paginate results
    paginated_data = paginate_results(sample_types, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/sample-types', methods=['POST'])
@token_required
def create_sample_type():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['type_name', 'type_code']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    sample_types = read_data('sample_types.json')

    # Check if code already exists
    if any(st.get('type_code') == data['type_code'].upper() for st in sample_types):
        return jsonify({'message': 'Sample type code already exists'}), 400

    # Generate new sample type ID
    new_id = 1
    if sample_types:
        new_id = max(st['id'] for st in sample_types) + 1

    # Create new sample type
    new_sample_type = {
        'id': new_id,
        'type_name': data['type_name'],
        'type_code': data['type_code'].upper(),
        'description': data.get('description', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    sample_types.append(new_sample_type)
    write_data('sample_types.json', sample_types)

    return jsonify(new_sample_type), 201

# Test Parameter Management Routes
@admin_bp.route('/api/admin/test-parameters', methods=['GET'])
@token_required
def get_test_parameters():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    test_parameters = read_data('test_parameters.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    test_parameters = sorted(test_parameters, key=lambda x: x.get('created_at', ''), reverse=True)

    # Paginate results
    paginated_data = paginate_results(test_parameters, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/test-parameters', methods=['POST'])
@token_required
def create_test_parameter():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['parameter_name', 'parameter_code']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    test_parameters = read_data('test_parameters.json')

    # Check if code already exists
    if any(tp.get('parameter_code') == data['parameter_code'].upper() for tp in test_parameters):
        return jsonify({'message': 'Parameter code already exists'}), 400

    # Generate new parameter ID
    new_id = 1
    if test_parameters:
        new_id = max(tp['id'] for tp in test_parameters) + 1

    # Create new test parameter
    new_parameter = {
        'id': new_id,
        'parameter_name': data['parameter_name'],
        'parameter_code': data['parameter_code'].upper(),
        'unit': data.get('unit', ''),
        'reference_range': data.get('reference_range', ''),
        'method': data.get('method', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    test_parameters.append(new_parameter)
    write_data('test_parameters.json', test_parameters)

    return jsonify(new_parameter), 201

# Department Management Routes
@admin_bp.route('/api/admin/departments', methods=['GET'])
@token_required
def get_departments():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    departments = read_data('departments.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    departments = sorted(departments, key=lambda x: x.get('created_at', ''), reverse=True)

    # Paginate results
    paginated_data = paginate_results(departments, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/departments', methods=['POST'])
@token_required
def create_department():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['department_name', 'department_code']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    departments = read_data('departments.json')

    # Check if code already exists
    if any(d.get('department_code') == data['department_code'].upper() for d in departments):
        return jsonify({'message': 'Department code already exists'}), 400

    # Generate new department ID
    new_id = 1
    if departments:
        new_id = max(d['id'] for d in departments) + 1

    # Create new department
    new_department = {
        'id': new_id,
        'department_name': data['department_name'],
        'department_code': data['department_code'].upper(),
        'description': data.get('description', ''),
        'head_of_department': data.get('head_of_department', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    departments.append(new_department)
    write_data('departments.json', departments)

    return jsonify(new_department), 201

# Payment Method Management Routes
@admin_bp.route('/api/admin/payment-methods', methods=['GET'])
@token_required
def get_payment_methods():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    payment_methods = read_data('payment_methods.json')

    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('limit', 20, type=int)

    # Sort by created_at (newest first)
    payment_methods = sorted(payment_methods, key=lambda x: x.get('created_at', ''), reverse=True)

    # Paginate results
    paginated_data = paginate_results(payment_methods, page, per_page)

    return jsonify(paginated_data)

@admin_bp.route('/api/admin/payment-methods', methods=['POST'])
@token_required
def create_payment_method():
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Validate required fields
    required_fields = ['method_name', 'method_code']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    payment_methods = read_data('payment_methods.json')

    # Check if code already exists
    if any(pm.get('method_code') == data['method_code'].upper() for pm in payment_methods):
        return jsonify({'message': 'Payment method code already exists'}), 400

    # Generate new payment method ID
    new_id = 1
    if payment_methods:
        new_id = max(pm['id'] for pm in payment_methods) + 1

    # Create new payment method
    new_payment_method = {
        'id': new_id,
        'method_name': data['method_name'],
        'method_code': data['method_code'].upper(),
        'description': data.get('description', ''),
        'is_online': data.get('is_online', False),
        'processing_fee': data.get('processing_fee', 0),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    payment_methods.append(new_payment_method)
    write_data('payment_methods.json', payment_methods)

    return jsonify(new_payment_method), 201

# Generic Master Data Routes (for frontend compatibility)
@admin_bp.route('/api/admin/master-data/<category>', methods=['POST'])
@token_required
def add_master_data_item(category):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Route to specific endpoints based on category
    if category == 'testCategories':
        return create_test_category_generic(data)
    elif category == 'testParameters':
        return create_test_parameter_generic(data)
    elif category == 'sampleTypes':
        return create_sample_type_generic(data)
    elif category == 'departments':
        return create_department_generic(data)
    elif category == 'paymentMethods':
        return create_payment_method_generic(data)
    elif category == 'containers':
        return create_container_generic(data)
    elif category == 'instruments':
        return create_instrument_generic(data)
    elif category == 'reagents':
        return create_reagent_generic(data)
    elif category == 'suppliers':
        return create_supplier_generic(data)
    elif category == 'units':
        return create_unit_generic(data)
    elif category == 'testMethods':
        return create_test_method_generic(data)
    elif category == 'patients':
        return create_patient_generic(data)
    elif category == 'profileMaster':
        return create_profile_master_generic(data)
    elif category == 'methodMaster':
        return create_method_master_generic(data)
    elif category == 'antibioticMaster':
        return create_antibiotic_master_generic(data)
    elif category == 'organismMaster':
        return create_organism_master_generic(data)
    elif category == 'unitOfMeasurement':
        return create_unit_of_measurement_generic(data)
    elif category == 'specimenMaster':
        return create_specimen_master_generic(data)
    elif category == 'organismVsAntibiotic':
        return create_organism_vs_antibiotic_generic(data)
    elif category == 'containerMaster':
        return create_container_master_generic(data)
    elif category == 'mainDepartmentMaster':
        return create_main_department_master_generic(data)
    elif category == 'departmentSettings':
        return create_department_settings_generic(data)
    elif category == 'authorizationSettings':
        return create_authorization_settings_generic(data)
    elif category == 'printOrder':
        return create_print_order_generic(data)
    elif category == 'testMaster':
        return create_test_master_generic(data)
    elif category == 'subTestMaster':
        return create_sub_test_master_generic(data)
    elif category == 'profileData':
        return create_profile_data_generic(data)
    else:
        return jsonify({'message': 'Invalid category'}), 400

@admin_bp.route('/api/admin/master-data/<category>/<int:item_id>', methods=['PUT'])
@token_required
def update_master_data_item(category, item_id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Route to specific endpoints based on category
    if category == 'testCategories':
        return update_test_category_generic(item_id, data)
    elif category == 'testParameters':
        return update_test_parameter_generic(item_id, data)
    elif category == 'sampleTypes':
        return update_sample_type_generic(item_id, data)
    elif category == 'departments':
        return update_department_generic(item_id, data)
    elif category == 'paymentMethods':
        return update_payment_method_generic(item_id, data)
    elif category == 'containers':
        return update_container_generic(item_id, data)
    elif category == 'instruments':
        return update_instrument_generic(item_id, data)
    elif category == 'reagents':
        return update_reagent_generic(item_id, data)
    elif category == 'suppliers':
        return update_supplier_generic(item_id, data)
    elif category == 'units':
        return update_unit_generic(item_id, data)
    elif category == 'testMethods':
        return update_test_method_generic(item_id, data)
    elif category == 'patients':
        return update_patient_generic(item_id, data)
    elif category == 'profileMaster':
        return update_profile_master_generic(item_id, data)
    elif category == 'methodMaster':
        return update_method_master_generic(item_id, data)
    elif category == 'antibioticMaster':
        return update_antibiotic_master_generic(item_id, data)
    elif category == 'organismMaster':
        return update_organism_master_generic(item_id, data)
    elif category == 'unitOfMeasurement':
        return update_unit_of_measurement_generic(item_id, data)
    elif category == 'specimenMaster':
        return update_specimen_master_generic(item_id, data)
    elif category == 'organismVsAntibiotic':
        return update_organism_vs_antibiotic_generic(item_id, data)
    elif category == 'containerMaster':
        return update_container_master_generic(item_id, data)
    elif category == 'mainDepartmentMaster':
        return update_main_department_master_generic(item_id, data)
    elif category == 'departmentSettings':
        return update_department_settings_generic(item_id, data)
    elif category == 'authorizationSettings':
        return update_authorization_settings_generic(item_id, data)
    elif category == 'printOrder':
        return update_print_order_generic(item_id, data)
    elif category == 'testMaster':
        return update_test_master_generic(item_id, data)
    elif category == 'subTestMaster':
        return update_sub_test_master_generic(item_id, data)
    elif category == 'profileData':
        return update_profile_data_generic(item_id, data)
    else:
        return jsonify({'message': 'Invalid category'}), 400

@admin_bp.route('/api/admin/master-data/<category>/<int:item_id>', methods=['DELETE'])
@token_required
def delete_master_data_item(category, item_id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    # Route to specific endpoints based on category
    if category == 'testCategories':
        return delete_test_category_generic(item_id)
    elif category == 'testParameters':
        return delete_test_parameter_generic(item_id)
    elif category == 'sampleTypes':
        return delete_sample_type_generic(item_id)
    elif category == 'departments':
        return delete_department_generic(item_id)
    elif category == 'paymentMethods':
        return delete_payment_method_generic(item_id)
    elif category == 'containers':
        return delete_container_generic(item_id)
    elif category == 'instruments':
        return delete_instrument_generic(item_id)
    elif category == 'reagents':
        return delete_reagent_generic(item_id)
    elif category == 'suppliers':
        return delete_supplier_generic(item_id)
    elif category == 'units':
        return delete_unit_generic(item_id)
    elif category == 'testMethods':
        return delete_test_method_generic(item_id)
    elif category == 'patients':
        return delete_patient_generic(item_id)
    elif category == 'profileMaster':
        return delete_profile_master_generic(item_id)
    elif category == 'methodMaster':
        return delete_method_master_generic(item_id)
    elif category == 'antibioticMaster':
        return delete_antibiotic_master_generic(item_id)
    elif category == 'organismMaster':
        return delete_organism_master_generic(item_id)
    elif category == 'unitOfMeasurement':
        return delete_unit_of_measurement_generic(item_id)
    elif category == 'specimenMaster':
        return delete_specimen_master_generic(item_id)
    elif category == 'organismVsAntibiotic':
        return delete_organism_vs_antibiotic_generic(item_id)
    elif category == 'containerMaster':
        return delete_container_master_generic(item_id)
    elif category == 'mainDepartmentMaster':
        return delete_main_department_master_generic(item_id)
    elif category == 'departmentSettings':
        return delete_department_settings_generic(item_id)
    elif category == 'authorizationSettings':
        return delete_authorization_settings_generic(item_id)
    elif category == 'printOrder':
        return delete_print_order_generic(item_id)
    elif category == 'testMaster':
        return delete_test_master_generic(item_id)
    elif category == 'subTestMaster':
        return delete_sub_test_master_generic(item_id)
    elif category == 'profileData':
        return delete_profile_data_generic(item_id)
    else:
        return jsonify({'message': 'Invalid category'}), 400

# Generic helper functions for each category
def create_test_category_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    categories = read_data('test_categories.json')

    # Generate code from name if not provided
    code = data.get('code', data['name'].upper().replace(' ', '_'))

    # Check if code already exists
    if any(c.get('code') == code for c in categories):
        return jsonify({'message': 'Category code already exists'}), 400

    # Generate new category ID
    new_id = 1
    if categories:
        new_id = max(c['id'] for c in categories) + 1

    # Create new test category
    new_category = {
        'id': new_id,
        'name': data['name'],
        'code': code,
        'description': data.get('description', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    categories.append(new_category)
    write_data('test_categories.json', categories)

    return jsonify(new_category), 201

def create_test_parameter_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    test_parameters = read_data('test_parameters.json')

    # Generate code from name if not provided
    code = data.get('parameter_code', data['name'].upper().replace(' ', '_'))

    # Check if code already exists
    if any(tp.get('parameter_code') == code for tp in test_parameters):
        return jsonify({'message': 'Parameter code already exists'}), 400

    # Generate new parameter ID
    new_id = 1
    if test_parameters:
        new_id = max(tp['id'] for tp in test_parameters) + 1

    # Create new test parameter
    new_parameter = {
        'id': new_id,
        'parameter_name': data['name'],
        'parameter_code': code,
        'unit': data.get('unit', ''),
        'reference_range': data.get('reference_range', ''),
        'method': data.get('method', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    test_parameters.append(new_parameter)
    write_data('test_parameters.json', test_parameters)

    return jsonify(new_parameter), 201

def create_sample_type_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    sample_types = read_data('sample_types.json')

    # Generate code from name if not provided
    code = data.get('type_code', data['name'].upper().replace(' ', '_'))

    # Check if code already exists
    if any(st.get('type_code') == code for st in sample_types):
        return jsonify({'message': 'Sample type code already exists'}), 400

    # Generate new sample type ID
    new_id = 1
    if sample_types:
        new_id = max(st['id'] for st in sample_types) + 1

    # Create new sample type
    new_sample_type = {
        'id': new_id,
        'type_name': data['name'],
        'type_code': code,
        'description': data.get('description', ''),
        'storage_instructions': data.get('storage_instructions', ''),
        'validity_days': data.get('validity_days', 7),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    sample_types.append(new_sample_type)
    write_data('sample_types.json', sample_types)

    return jsonify(new_sample_type), 201

def create_department_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    departments = read_data('departments.json')

    # Generate code from name if not provided
    code = data.get('department_code', data['name'].upper().replace(' ', '_'))

    # Check if code already exists
    if any(d.get('department_code') == code for d in departments):
        return jsonify({'message': 'Department code already exists'}), 400

    # Generate new department ID
    new_id = 1
    if departments:
        new_id = max(d['id'] for d in departments) + 1

    # Create new department
    new_department = {
        'id': new_id,
        'department_name': data['name'],
        'department_code': code,
        'description': data.get('description', ''),
        'head_of_department': data.get('head_of_department', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    departments.append(new_department)
    write_data('departments.json', departments)

    return jsonify(new_department), 201

def create_payment_method_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    payment_methods = read_data('payment_methods.json')

    # Generate code from name if not provided
    code = data.get('method_code', data['name'].upper().replace(' ', '_'))

    # Check if code already exists
    if any(pm.get('method_code') == code for pm in payment_methods):
        return jsonify({'message': 'Payment method code already exists'}), 400

    # Generate new payment method ID
    new_id = 1
    if payment_methods:
        new_id = max(pm['id'] for pm in payment_methods) + 1

    # Create new payment method
    new_payment_method = {
        'id': new_id,
        'method_name': data['name'],
        'method_code': code,
        'description': data.get('description', ''),
        'is_online': data.get('is_online', False),
        'processing_fee': data.get('processing_fee', 0),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    payment_methods.append(new_payment_method)
    write_data('payment_methods.json', payment_methods)

    return jsonify(new_payment_method), 201

def create_container_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    containers = read_data('containers.json')

    # Generate new container ID
    new_id = 1
    if containers:
        new_id = max(c['id'] for c in containers) + 1

    # Create new container
    new_container = {
        'id': new_id,
        'name': data['name'],
        'type': data.get('type', ''),
        'volume': data.get('volume', ''),
        'unit': data.get('unit', ''),
        'color': data.get('color', ''),
        'additive': data.get('additive', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    containers.append(new_container)
    write_data('containers.json', containers)

    return jsonify(new_container), 201

def create_instrument_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    instruments = read_data('instruments.json')

    # Generate new instrument ID
    new_id = 1
    if instruments:
        new_id = max(i['id'] for i in instruments) + 1

    # Create new instrument
    new_instrument = {
        'id': new_id,
        'name': data['name'],
        'model': data.get('model', ''),
        'manufacturer': data.get('manufacturer', ''),
        'serial_number': data.get('serial_number', ''),
        'installation_date': data.get('installation_date', ''),
        'calibration_due': data.get('calibration_due', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    instruments.append(new_instrument)
    write_data('instruments.json', instruments)

    return jsonify(new_instrument), 201

def create_reagent_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    reagents = read_data('reagents.json')

    # Generate new reagent ID
    new_id = 1
    if reagents:
        new_id = max(r['id'] for r in reagents) + 1

    # Create new reagent
    new_reagent = {
        'id': new_id,
        'name': data['name'],
        'lot_number': data.get('lot_number', ''),
        'expiry_date': data.get('expiry_date', ''),
        'manufacturer': data.get('manufacturer', ''),
        'storage_temperature': data.get('storage_temperature', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    reagents.append(new_reagent)
    write_data('reagents.json', reagents)

    return jsonify(new_reagent), 201

def create_supplier_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    suppliers = read_data('suppliers.json')

    # Generate new supplier ID
    new_id = 1
    if suppliers:
        new_id = max(s['id'] for s in suppliers) + 1

    # Create new supplier
    new_supplier = {
        'id': new_id,
        'name': data['name'],
        'contact_person': data.get('contact_person', ''),
        'email': data.get('email', ''),
        'phone': data.get('phone', ''),
        'address': data.get('address', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    suppliers.append(new_supplier)
    write_data('suppliers.json', suppliers)

    return jsonify(new_supplier), 201

def create_unit_generic(data):
    # Validate required fields
    required_fields = ['name', 'symbol']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    units = read_data('units.json')

    # Check if symbol already exists
    if any(u.get('symbol') == data['symbol'] for u in units):
        return jsonify({'message': 'Unit symbol already exists'}), 400

    # Generate new unit ID
    new_id = 1
    if units:
        new_id = max(u['id'] for u in units) + 1

    # Create new unit
    new_unit = {
        'id': new_id,
        'name': data['name'],
        'symbol': data['symbol'],
        'type': data.get('type', ''),
        'conversion_factor': data.get('conversion_factor', 1),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    units.append(new_unit)
    write_data('units.json', units)

    return jsonify(new_unit), 201

def create_test_method_generic(data):
    # Validate required fields
    required_fields = ['name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    test_methods = read_data('test_methods.json')

    # Generate new test method ID
    new_id = 1
    if test_methods:
        new_id = max(tm['id'] for tm in test_methods) + 1

    # Create new test method
    new_test_method = {
        'id': new_id,
        'name': data['name'],
        'description': data.get('description', ''),
        'principle': data.get('principle', ''),
        'procedure': data.get('procedure', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    test_methods.append(new_test_method)
    write_data('test_methods.json', test_methods)

    return jsonify(new_test_method), 201

# Generic update functions
def update_test_category_generic(item_id, data):
    categories = read_data('test_categories.json')
    category_index = next((i for i, c in enumerate(categories) if c['id'] == item_id), None)

    if category_index is None:
        return jsonify({'message': 'Test category not found'}), 404

    category = categories[category_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'code':
                category[key] = value.upper()
            else:
                category[key] = value

    category['updated_at'] = datetime.now().isoformat()
    write_data('test_categories.json', categories)
    return jsonify(category)

def update_test_parameter_generic(item_id, data):
    test_parameters = read_data('test_parameters.json')
    parameter_index = next((i for i, tp in enumerate(test_parameters) if tp['id'] == item_id), None)

    if parameter_index is None:
        return jsonify({'message': 'Test parameter not found'}), 404

    parameter = test_parameters[parameter_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'parameter_code':
                parameter[key] = value.upper()
            else:
                parameter[key] = value

    parameter['updated_at'] = datetime.now().isoformat()
    write_data('test_parameters.json', test_parameters)
    return jsonify(parameter)

def update_sample_type_generic(item_id, data):
    sample_types = read_data('sample_types.json')
    sample_type_index = next((i for i, st in enumerate(sample_types) if st['id'] == item_id), None)

    if sample_type_index is None:
        return jsonify({'message': 'Sample type not found'}), 404

    sample_type = sample_types[sample_type_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'type_code':
                sample_type[key] = value.upper()
            else:
                sample_type[key] = value

    sample_type['updated_at'] = datetime.now().isoformat()
    write_data('sample_types.json', sample_types)
    return jsonify(sample_type)

def update_department_generic(item_id, data):
    departments = read_data('departments.json')
    department_index = next((i for i, d in enumerate(departments) if d['id'] == item_id), None)

    if department_index is None:
        return jsonify({'message': 'Department not found'}), 404

    department = departments[department_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'department_code':
                department[key] = value.upper()
            else:
                department[key] = value

    department['updated_at'] = datetime.now().isoformat()
    write_data('departments.json', departments)
    return jsonify(department)

def update_payment_method_generic(item_id, data):
    payment_methods = read_data('payment_methods.json')
    payment_method_index = next((i for i, pm in enumerate(payment_methods) if pm['id'] == item_id), None)

    if payment_method_index is None:
        return jsonify({'message': 'Payment method not found'}), 404

    payment_method = payment_methods[payment_method_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'method_code':
                payment_method[key] = value.upper()
            else:
                payment_method[key] = value

    payment_method['updated_at'] = datetime.now().isoformat()
    write_data('payment_methods.json', payment_methods)
    return jsonify(payment_method)

def update_container_generic(item_id, data):
    containers = read_data('containers.json')
    container_index = next((i for i, c in enumerate(containers) if c['id'] == item_id), None)

    if container_index is None:
        return jsonify({'message': 'Container not found'}), 404

    container = containers[container_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            container[key] = value

    container['updated_at'] = datetime.now().isoformat()
    write_data('containers.json', containers)
    return jsonify(container)

def update_instrument_generic(item_id, data):
    instruments = read_data('instruments.json')
    instrument_index = next((i for i, inst in enumerate(instruments) if inst['id'] == item_id), None)

    if instrument_index is None:
        return jsonify({'message': 'Instrument not found'}), 404

    instrument = instruments[instrument_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            instrument[key] = value

    instrument['updated_at'] = datetime.now().isoformat()
    write_data('instruments.json', instruments)
    return jsonify(instrument)

def update_reagent_generic(item_id, data):
    reagents = read_data('reagents.json')
    reagent_index = next((i for i, r in enumerate(reagents) if r['id'] == item_id), None)

    if reagent_index is None:
        return jsonify({'message': 'Reagent not found'}), 404

    reagent = reagents[reagent_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            reagent[key] = value

    reagent['updated_at'] = datetime.now().isoformat()
    write_data('reagents.json', reagents)
    return jsonify(reagent)

def update_supplier_generic(item_id, data):
    suppliers = read_data('suppliers.json')
    supplier_index = next((i for i, s in enumerate(suppliers) if s['id'] == item_id), None)

    if supplier_index is None:
        return jsonify({'message': 'Supplier not found'}), 404

    supplier = suppliers[supplier_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            supplier[key] = value

    supplier['updated_at'] = datetime.now().isoformat()
    write_data('suppliers.json', suppliers)
    return jsonify(supplier)

def update_unit_generic(item_id, data):
    units = read_data('units.json')
    unit_index = next((i for i, u in enumerate(units) if u['id'] == item_id), None)

    if unit_index is None:
        return jsonify({'message': 'Unit not found'}), 404

    unit = units[unit_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            unit[key] = value

    unit['updated_at'] = datetime.now().isoformat()
    write_data('units.json', units)
    return jsonify(unit)

def update_test_method_generic(item_id, data):
    test_methods = read_data('test_methods.json')
    method_index = next((i for i, tm in enumerate(test_methods) if tm['id'] == item_id), None)

    if method_index is None:
        return jsonify({'message': 'Test method not found'}), 404

    test_method = test_methods[method_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            test_method[key] = value

    test_method['updated_at'] = datetime.now().isoformat()
    write_data('test_methods.json', test_methods)
    return jsonify(test_method)

# Generic delete functions
def delete_test_category_generic(item_id):
    categories = read_data('test_categories.json')
    category_index = next((i for i, c in enumerate(categories) if c['id'] == item_id), None)

    if category_index is None:
        return jsonify({'message': 'Test category not found'}), 404

    categories.pop(category_index)
    write_data('test_categories.json', categories)
    return jsonify({'message': 'Test category deleted successfully'})

def delete_test_parameter_generic(item_id):
    test_parameters = read_data('test_parameters.json')
    parameter_index = next((i for i, tp in enumerate(test_parameters) if tp['id'] == item_id), None)

    if parameter_index is None:
        return jsonify({'message': 'Test parameter not found'}), 404

    test_parameters.pop(parameter_index)
    write_data('test_parameters.json', test_parameters)
    return jsonify({'message': 'Test parameter deleted successfully'})

def delete_sample_type_generic(item_id):
    sample_types = read_data('sample_types.json')
    sample_type_index = next((i for i, st in enumerate(sample_types) if st['id'] == item_id), None)

    if sample_type_index is None:
        return jsonify({'message': 'Sample type not found'}), 404

    sample_types.pop(sample_type_index)
    write_data('sample_types.json', sample_types)
    return jsonify({'message': 'Sample type deleted successfully'})

def delete_department_generic(item_id):
    departments = read_data('departments.json')
    department_index = next((i for i, d in enumerate(departments) if d['id'] == item_id), None)

    if department_index is None:
        return jsonify({'message': 'Department not found'}), 404

    departments.pop(department_index)
    write_data('departments.json', departments)
    return jsonify({'message': 'Department deleted successfully'})

def delete_payment_method_generic(item_id):
    payment_methods = read_data('payment_methods.json')
    payment_method_index = next((i for i, pm in enumerate(payment_methods) if pm['id'] == item_id), None)

    if payment_method_index is None:
        return jsonify({'message': 'Payment method not found'}), 404

    payment_methods.pop(payment_method_index)
    write_data('payment_methods.json', payment_methods)
    return jsonify({'message': 'Payment method deleted successfully'})

def delete_container_generic(item_id):
    containers = read_data('containers.json')
    container_index = next((i for i, c in enumerate(containers) if c['id'] == item_id), None)

    if container_index is None:
        return jsonify({'message': 'Container not found'}), 404

    containers.pop(container_index)
    write_data('containers.json', containers)
    return jsonify({'message': 'Container deleted successfully'})

def delete_instrument_generic(item_id):
    instruments = read_data('instruments.json')
    instrument_index = next((i for i, inst in enumerate(instruments) if inst['id'] == item_id), None)

    if instrument_index is None:
        return jsonify({'message': 'Instrument not found'}), 404

    instruments.pop(instrument_index)
    write_data('instruments.json', instruments)
    return jsonify({'message': 'Instrument deleted successfully'})

def delete_reagent_generic(item_id):
    reagents = read_data('reagents.json')
    reagent_index = next((i for i, r in enumerate(reagents) if r['id'] == item_id), None)

    if reagent_index is None:
        return jsonify({'message': 'Reagent not found'}), 404

    reagents.pop(reagent_index)
    write_data('reagents.json', reagents)
    return jsonify({'message': 'Reagent deleted successfully'})

def delete_supplier_generic(item_id):
    suppliers = read_data('suppliers.json')
    supplier_index = next((i for i, s in enumerate(suppliers) if s['id'] == item_id), None)

    if supplier_index is None:
        return jsonify({'message': 'Supplier not found'}), 404

    suppliers.pop(supplier_index)
    write_data('suppliers.json', suppliers)
    return jsonify({'message': 'Supplier deleted successfully'})

def delete_unit_generic(item_id):
    units = read_data('units.json')
    unit_index = next((i for i, u in enumerate(units) if u['id'] == item_id), None)

    if unit_index is None:
        return jsonify({'message': 'Unit not found'}), 404

    units.pop(unit_index)
    write_data('units.json', units)
    return jsonify({'message': 'Unit deleted successfully'})

def delete_test_method_generic(item_id):
    test_methods = read_data('test_methods.json')
    method_index = next((i for i, tm in enumerate(test_methods) if tm['id'] == item_id), None)

    if method_index is None:
        return jsonify({'message': 'Test method not found'}), 404

    test_methods.pop(method_index)
    write_data('test_methods.json', test_methods)
    return jsonify({'message': 'Test method deleted successfully'})

# Excel Import/Export endpoints
@admin_bp.route('/master-data/import', methods=['POST'])
@token_required
def import_master_data_old():
    """Import master data from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'message': 'No file provided'}), 400

        file = request.files['file']
        category = request.form.get('category')

        if not file or file.filename == '':
            return jsonify({'message': 'No file selected'}), 400

        if not category:
            return jsonify({'message': 'Category not specified'}), 400

        # Save uploaded file temporarily
        import tempfile
        import os

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)

            try:
                # Process the Excel file
                result = process_excel_import(tmp_file.name, category)
                return jsonify(result)
            finally:
                # Clean up temporary file
                os.unlink(tmp_file.name)

    except Exception as e:
        return jsonify({'message': f'Import failed: {str(e)}'}), 500

@admin_bp.route('/master-data/export', methods=['GET'])
@token_required
def export_master_data_old():
    """Export master data to Excel file"""
    try:
        category = request.args.get('category')

        if not category:
            return jsonify({'message': 'Category not specified'}), 400

        # Generate Excel file
        excel_file = generate_excel_export(category)

        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f'{category}_master_data.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'message': f'Export failed: {str(e)}'}), 500

@admin_bp.route('/master-data/template', methods=['GET'])
@token_required
def download_template_v2():
    """Download Excel template for master data import"""
    try:
        category = request.args.get('category')

        if not category:
            return jsonify({'message': 'Category not specified'}), 400

        # Generate template file
        template_file = generate_excel_template(category)

        return send_file(
            template_file,
            as_attachment=True,
            download_name=f'{category}_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'message': f'Template download failed: {str(e)}'}), 500

# Helper functions for Excel processing
def process_excel_import(file_path, category):
    """Process Excel file import for master data"""
    try:
        # Try to import pandas, if not available, return error
        try:
            import pandas as pd
        except ImportError:
            return {
                'success_count': 0,
                'error_count': 1,
                'total_rows': 0,
                'errors': [{'row': 0, 'field': 'system', 'message': 'pandas library not installed. Please install pandas to use Excel import functionality.'}]
            }

        # Read Excel file
        df = pd.read_excel(file_path)

        if df.empty:
            return {
                'success_count': 0,
                'error_count': 0,
                'total_rows': 0,
                'errors': []
            }

        # Get existing data
        filename = get_filename_for_category(category)
        existing_data = read_data(filename)

        # Get next ID
        next_id = 1
        if existing_data:
            next_id = max(item.get('id', 0) for item in existing_data) + 1

        success_count = 0
        error_count = 0
        errors = []

        # Process each row
        for index, row in df.iterrows():
            try:
                # Validate and convert row data
                item_data = validate_and_convert_row(row, category, index + 2)  # +2 for header and 0-based index

                if item_data:
                    # Add metadata
                    item_data['id'] = next_id
                    item_data['created_at'] = datetime.now().isoformat()
                    item_data['updated_at'] = datetime.now().isoformat()
                    item_data['created_by'] = 1  # Default admin user

                    existing_data.append(item_data)
                    next_id += 1
                    success_count += 1
                else:
                    error_count += 1

            except Exception as e:
                error_count += 1
                errors.append({
                    'row': index + 2,
                    'field': 'general',
                    'message': str(e)
                })

        # Save updated data
        if success_count > 0:
            write_data(filename, existing_data)

        return {
            'success_count': success_count,
            'error_count': error_count,
            'total_rows': len(df),
            'errors': errors
        }

    except Exception as e:
        return {
            'success_count': 0,
            'error_count': 1,
            'total_rows': 0,
            'errors': [{'row': 0, 'field': 'system', 'message': f'Failed to process Excel file: {str(e)}'}]
        }

def validate_and_convert_row(row, category, row_number):
    """Validate and convert a row of data based on category"""
    errors = []

    # Convert pandas Series to dict and handle NaN values
    row_dict = {}
    for key, value in row.items():
        if pd.isna(value):
            row_dict[key] = ''
        else:
            row_dict[key] = str(value).strip()

    # Category-specific validation
    if category == 'testCategories':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'description': row_dict.get('description', ''),
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'testParameters':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'unit': row_dict.get('unit', ''),
            'reference_range': row_dict.get('reference_range', ''),
            'category_id': int(row_dict.get('category_id', 1)) if row_dict.get('category_id') else 1,
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'sampleTypes':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'description': row_dict.get('description', ''),
            'storage_instructions': row_dict.get('storage_instructions', ''),
            'validity_days': int(row_dict.get('validity_days', 7)) if row_dict.get('validity_days') else 7,
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'departments':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'description': row_dict.get('description', ''),
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'paymentMethods':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'description': row_dict.get('description', ''),
            'is_online': str(row_dict.get('is_online', 'false')).lower() in ['true', '1', 'yes', 'online'],
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'containers':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'type': row_dict.get('type', ''),
            'volume': row_dict.get('volume', ''),
            'unit': row_dict.get('unit', ''),
            'color': row_dict.get('color', ''),
            'additive': row_dict.get('additive', ''),
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'instruments':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'model': row_dict.get('model', ''),
            'manufacturer': row_dict.get('manufacturer', ''),
            'serial_number': row_dict.get('serial_number', ''),
            'installation_date': row_dict.get('installation_date', ''),
            'calibration_due': row_dict.get('calibration_due', ''),
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'reagents':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'lot_number': row_dict.get('lot_number', ''),
            'expiry_date': row_dict.get('expiry_date', ''),
            'manufacturer': row_dict.get('manufacturer', ''),
            'storage_temperature': row_dict.get('storage_temperature', ''),
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'suppliers':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'contact_person': row_dict.get('contact_person', ''),
            'email': row_dict.get('email', ''),
            'phone': row_dict.get('phone', ''),
            'address': row_dict.get('address', ''),
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'units':
        if not row_dict.get('name') or not row_dict.get('symbol'):
            errors.append({'row': row_number, 'field': 'name/symbol', 'message': 'Name and symbol are required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'symbol': row_dict.get('symbol', ''),
            'type': row_dict.get('type', ''),
            'conversion_factor': float(row_dict.get('conversion_factor', 1)) if row_dict.get('conversion_factor') else 1,
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    elif category == 'testMethods':
        if not row_dict.get('name'):
            errors.append({'row': row_number, 'field': 'name', 'message': 'Name is required'})
            return None

        return {
            'name': row_dict.get('name', ''),
            'description': row_dict.get('description', ''),
            'principle': row_dict.get('principle', ''),
            'procedure': row_dict.get('procedure', ''),
            'is_active': str(row_dict.get('is_active', 'true')).lower() in ['true', '1', 'yes', 'active']
        }

    return None

def generate_excel_export(category):
    """Generate Excel file for export"""
    try:
        import pandas as pd
    except ImportError:
        raise Exception("pandas library not installed. Please install pandas to use Excel export functionality.")

    # Get data
    filename = get_filename_for_category(category)
    data = read_data(filename)

    if not data:
        # Create empty DataFrame with headers
        df = pd.DataFrame(columns=get_headers_for_category(category))
    else:
        # Convert data to DataFrame
        df = pd.DataFrame(data)

        # Select only relevant columns
        headers = get_headers_for_category(category)
        df = df.reindex(columns=headers, fill_value='')

    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=category, index=False)

    output.seek(0)
    return output

def generate_excel_template(category):
    """Generate Excel template for import"""
    try:
        import pandas as pd
    except ImportError:
        raise Exception("pandas library not installed. Please install pandas to use Excel template functionality.")

    # Get headers for category
    headers = get_headers_for_category(category)

    # Create sample data
    sample_data = get_sample_data_for_category(category)

    # Create DataFrame
    df = pd.DataFrame([sample_data], columns=headers)

    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f'{category}_template', index=False)

    output.seek(0)
    return output

def get_filename_for_category(category):
    """Get filename for category"""
    filename_map = {
        'testCategories': 'test_categories.json',
        'testParameters': 'test_parameters.json',
        'sampleTypes': 'sample_types.json',
        'departments': 'departments.json',
        'paymentMethods': 'payment_methods.json',
        'containers': 'containers.json',
        'instruments': 'instruments.json',
        'reagents': 'reagents.json',
        'suppliers': 'suppliers.json',
        'units': 'units.json',
        'testMethods': 'test_methods.json'
    }
    return filename_map.get(category, f'{category}.json')

def get_headers_for_category(category):
    """Get column headers for category"""
    headers_map = {
        'testCategories': ['name', 'description', 'is_active'],
        'testParameters': ['name', 'unit', 'reference_range', 'category_id', 'is_active'],
        'sampleTypes': ['name', 'description', 'storage_instructions', 'validity_days', 'is_active'],
        'departments': ['name', 'description', 'is_active'],
        'paymentMethods': ['name', 'description', 'is_online', 'is_active'],
        'containers': ['name', 'type', 'volume', 'unit', 'color', 'additive', 'is_active'],
        'instruments': ['name', 'model', 'manufacturer', 'serial_number', 'installation_date', 'calibration_due', 'is_active'],
        'reagents': ['name', 'lot_number', 'expiry_date', 'manufacturer', 'storage_temperature', 'is_active'],
        'suppliers': ['name', 'contact_person', 'email', 'phone', 'address', 'is_active'],
        'units': ['name', 'symbol', 'type', 'conversion_factor', 'is_active'],
        'testMethods': ['name', 'description', 'principle', 'procedure', 'is_active']
    }
    return headers_map.get(category, ['name', 'description', 'is_active'])

def get_sample_data_for_category(category):
    """Get sample data for category template"""
    sample_data_map = {
        'testCategories': {
            'name': 'Sample Category',
            'description': 'Sample description for test category',
            'is_active': 'true'
        },
        'testParameters': {
            'name': 'Sample Parameter',
            'unit': 'mg/dL',
            'reference_range': '70-100',
            'category_id': '1',
            'is_active': 'true'
        },
        'sampleTypes': {
            'name': 'Sample Type',
            'description': 'Sample description',
            'storage_instructions': 'Store at room temperature',
            'validity_days': '7',
            'is_active': 'true'
        },
        'departments': {
            'name': 'Sample Department',
            'description': 'Sample department description',
            'is_active': 'true'
        },
        'paymentMethods': {
            'name': 'Sample Payment Method',
            'description': 'Sample payment method description',
            'is_online': 'false',
            'is_active': 'true'
        },
        'containers': {
            'name': 'Sample Container',
            'type': 'Tube',
            'volume': '5',
            'unit': 'mL',
            'color': '#FF0000',
            'additive': 'EDTA',
            'is_active': 'true'
        },
        'instruments': {
            'name': 'Sample Instrument',
            'model': 'Model123',
            'manufacturer': 'Sample Manufacturer',
            'serial_number': 'SN123456',
            'installation_date': '2024-01-01',
            'calibration_due': '2024-12-31',
            'is_active': 'true'
        },
        'reagents': {
            'name': 'Sample Reagent',
            'lot_number': 'LOT123',
            'expiry_date': '2024-12-31',
            'manufacturer': 'Sample Manufacturer',
            'storage_temperature': '2-8°C',
            'is_active': 'true'
        },
        'suppliers': {
            'name': 'Sample Supplier',
            'contact_person': 'John Doe',
            'email': 'john@example.com',
            'phone': '+91-9876543210',
            'address': 'Sample Address',
            'is_active': 'true'
        },
        'units': {
            'name': 'Sample Unit',
            'symbol': 'mg/dL',
            'type': 'Concentration',
            'conversion_factor': '1.0',
            'is_active': 'true'
        },
        'testMethods': {
            'name': 'Sample Method',
            'description': 'Sample method description',
            'principle': 'Sample principle',
            'procedure': 'Sample procedure',
            'is_active': 'true'
        }
    }
    return sample_data_map.get(category, {'name': 'Sample', 'description': 'Sample description', 'is_active': 'true'})

@admin_bp.route('/api/admin/sample-types/<int:id>', methods=['PUT'])
@token_required
def update_sample_type(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    sample_types = read_data('sample_types.json')
    sample_type_index = next((i for i, st in enumerate(sample_types) if st['id'] == id), None)

    if sample_type_index is None:
        return jsonify({'message': 'Sample type not found'}), 404

    # Update sample type fields
    sample_type = sample_types[sample_type_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            if key == 'type_code':
                sample_type[key] = value.upper()
            else:
                sample_type[key] = value

    sample_type['updated_at'] = datetime.now().isoformat()

    # Save updated sample types
    write_data('sample_types.json', sample_types)

    return jsonify(sample_type)

@admin_bp.route('/api/admin/sample-types/<int:id>', methods=['DELETE'])
@token_required
def delete_sample_type(id):
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    sample_types = read_data('sample_types.json')
    sample_type_index = next((i for i, st in enumerate(sample_types) if st['id'] == id), None)

    if sample_type_index is None:
        return jsonify({'message': 'Sample type not found'}), 404

    # Delete sample type
    deleted_sample_type = sample_types.pop(sample_type_index)
    write_data('sample_types.json', sample_types)

    return jsonify({'message': 'Sample type deleted successfully'})

# Excel Import/Export Routes for Master Data
@admin_bp.route('/api/admin/master-data/import', methods=['POST'])
@token_required
def import_master_data():
    """Import master data from Excel file"""
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'message': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'message': 'No file selected'}), 400

        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'message': 'Invalid file format. Please upload an Excel file.'}), 400

        # Get category from form data
        category = request.form.get('category', '')
        if not category:
            return jsonify({'message': 'Category is required'}), 400

        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)

        try:
            # Read Excel file
            import pandas as pd
            df = pd.read_excel(temp_path)

            # Convert DataFrame to list of dictionaries
            records = df.to_dict('records')

            # Process records based on category
            processed_records = []
            existing_data = read_data(f'{category}.json')

            # Get next ID
            next_id = 1
            if existing_data:
                next_id = max(item.get('id', 0) for item in existing_data) + 1

            for record in records:
                # Clean the record (remove NaN values)
                clean_record = {}
                for key, value in record.items():
                    if pd.notna(value):
                        clean_record[key.lower().replace(' ', '_')] = value

                # Add metadata
                clean_record['id'] = next_id
                clean_record['is_active'] = clean_record.get('is_active', True)
                clean_record['created_at'] = datetime.now().isoformat()
                clean_record['updated_at'] = datetime.now().isoformat()
                clean_record['created_by'] = request.current_user.get('id', 1)

                processed_records.append(clean_record)
                next_id += 1

            # Append to existing data
            existing_data.extend(processed_records)

            # Save updated data
            write_data(f'{category}.json', existing_data)

            return jsonify({
                'message': f'Successfully imported {len(processed_records)} records',
                'imported_count': len(processed_records)
            })

        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)

    except Exception as e:
        return jsonify({'message': f'Import failed: {str(e)}'}), 500

@admin_bp.route('/api/admin/master-data/export/<category>', methods=['GET'])
@token_required
def export_master_data(category):
    """Export master data to Excel file"""
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        # Read data for the category
        data = read_data(f'{category}.json')

        if not data:
            return jsonify({'message': f'No data found for category: {category}'}), 404

        # Create DataFrame
        import pandas as pd
        df = pd.DataFrame(data)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=category, index=False)

        output.seek(0)

        # Return file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{category}_export.xlsx'
        )

    except Exception as e:
        return jsonify({'message': f'Export failed: {str(e)}'}), 500

@admin_bp.route('/api/admin/master-data/export', methods=['GET'])
@token_required
def export_all_master_data():
    """Export all master data to Excel file with multiple sheets"""
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        # List of all categories
        categories = [
            'testCategories', 'testParameters', 'sampleTypes', 'departments',
            'paymentMethods', 'containers', 'instruments', 'reagents',
            'suppliers', 'units', 'testMethods', 'patients', 'profileMaster',
            'methodMaster', 'antibioticMaster', 'organismMaster',
            'unitOfMeasurement', 'specimenMaster', 'organismVsAntibiotic',
            'containerMaster', 'mainDepartmentMaster', 'departmentSettings',
            'authorizationSettings', 'printOrder'
        ]

        # Create Excel file in memory
        import pandas as pd
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for category in categories:
                try:
                    # Convert camelCase to snake_case for file names
                    file_name = ''.join(['_' + c.lower() if c.isupper() else c for c in category]).lstrip('_')
                    data = read_data(f'{file_name}.json')

                    if data:
                        df = pd.DataFrame(data)
                        # Limit sheet name to 31 characters (Excel limit)
                        sheet_name = category[:31]
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception as e:
                    print(f"Error processing category {category}: {e}")
                    continue

        output.seek(0)

        # Return file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='master_data_export.xlsx'
        )

    except Exception as e:
        return jsonify({'message': f'Export failed: {str(e)}'}), 500

@admin_bp.route('/api/admin/master-data/template/<category>', methods=['GET'])
@token_required
def download_template(category):
    """Download Excel template for a specific category"""
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        # Define templates for each category
        templates = {
            'patients': ['his_no', 'patient_name', 'mobile', 'whatsapp_no', 'uid_no'],
            'profileMaster': ['code', 'procedure_code', 'test_profile', 'test_price'],
            'methodMaster': ['code', 'method'],
            'antibioticMaster': ['antibiotic_code', 'antibiotic_group', 'antibiotic_description', 'antibiotic_content', 'order'],
            'organismMaster': ['code', 'description', 'no_growth'],
            'unitOfMeasurement': ['code', 'description', 'technical', 'inventory'],
            'specimenMaster': ['code', 'specimen', 'container', 'disposable'],
            'organismVsAntibiotic': ['organism', 'antibiotic_group'],
            'containerMaster': ['code', 'description', 'short_name', 'color'],
            'mainDepartmentMaster': ['major_department', 'code', 'department', 'order', 'short_name', 'queue'],
            'departmentSettings': ['main', 'code', 'sub_name', 'service_time', 'room', 'order', 'dept_amt', 'short'],
            'authorizationSettings': ['main', 'code', 'sub_name', 'service_time', 'authorization', 'authorization_type', 'email_at', 'report_type'],
            'printOrder': ['item', 'order']
        }

        if category not in templates:
            return jsonify({'message': f'Template not available for category: {category}'}), 404

        # Create template DataFrame
        import pandas as pd
        template_data = {col: [''] for col in templates[category]}
        df = pd.DataFrame(template_data)

        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Template', index=False)

        output.seek(0)

        # Return file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{category}_template.xlsx'
        )

    except Exception as e:
        return jsonify({'message': f'Template generation failed: {str(e)}'}), 500

@admin_bp.route('/api/admin/master-data/bulk-import', methods=['POST'])
@token_required
def bulk_import_master_data():
    """Bulk import multiple categories from Excel file"""
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'message': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'message': 'No file selected'}), 400

        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'message': 'Invalid file format. Please upload an Excel file.'}), 400

        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)

        try:
            # Read Excel file with all sheets
            import pandas as pd
            excel_file = pd.ExcelFile(temp_path)

            results = {}
            total_imported = 0

            for sheet_name in excel_file.sheet_names:
                try:
                    # Read sheet
                    df = pd.read_excel(temp_path, sheet_name=sheet_name)

                    if df.empty:
                        continue

                    # Convert sheet name to file name
                    category = sheet_name.lower().replace(' ', '_')

                    # Convert DataFrame to list of dictionaries
                    records = df.to_dict('records')

                    # Process records
                    processed_records = []
                    existing_data = read_data(f'{category}.json')

                    # Get next ID
                    next_id = 1
                    if existing_data:
                        next_id = max(item.get('id', 0) for item in existing_data) + 1

                    for record in records:
                        # Clean the record (remove NaN values)
                        clean_record = {}
                        for key, value in record.items():
                            if pd.notna(value):
                                clean_record[key.lower().replace(' ', '_')] = value

                        # Add metadata
                        clean_record['id'] = next_id
                        clean_record['is_active'] = clean_record.get('is_active', True)
                        clean_record['created_at'] = datetime.now().isoformat()
                        clean_record['updated_at'] = datetime.now().isoformat()
                        clean_record['created_by'] = request.current_user.get('id', 1)

                        processed_records.append(clean_record)
                        next_id += 1

                    # Append to existing data
                    existing_data.extend(processed_records)

                    # Save updated data
                    write_data(f'{category}.json', existing_data)

                    results[sheet_name] = len(processed_records)
                    total_imported += len(processed_records)

                except Exception as e:
                    results[sheet_name] = f'Error: {str(e)}'

            return jsonify({
                'message': f'Bulk import completed. Total records imported: {total_imported}',
                'results': results,
                'total_imported': total_imported
            })

        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)

    except Exception as e:
        return jsonify({'message': f'Bulk import failed: {str(e)}'}), 500

# Technical Master Data Routes
@admin_bp.route('/api/admin/technical-master-data', methods=['GET'])
@token_required
def get_technical_master_data():
    """Get all technical master data"""
    try:
        # Initialize empty data structure
        technical_data = {
            'resultMaster': [],
            'parameterMaster': [],
            'referenceRanges': [],
            'calculationFormulas': [],
            'qualityControlRules': [],
            'instrumentMaster': [],
            'reagentMaster': [],
            'calibrationStandards': [],
            'referrerMaster': []
        }

        # Try to read existing data files
        try:
            technical_data['resultMaster'] = read_data('result_master.json')
        except:
            technical_data['resultMaster'] = []

        try:
            technical_data['parameterMaster'] = read_data('parameter_master.json')
        except:
            technical_data['parameterMaster'] = []

        try:
            technical_data['referenceRanges'] = read_data('reference_ranges.json')
        except:
            technical_data['referenceRanges'] = []

        try:
            technical_data['calculationFormulas'] = read_data('calculation_formulas.json')
        except:
            technical_data['calculationFormulas'] = []

        try:
            technical_data['qualityControlRules'] = read_data('quality_control_rules.json')
        except:
            technical_data['qualityControlRules'] = []

        try:
            technical_data['instrumentMaster'] = read_data('instrument_master.json')
        except:
            technical_data['instrumentMaster'] = []

        try:
            technical_data['reagentMaster'] = read_data('reagent_master.json')
        except:
            technical_data['reagentMaster'] = []

        try:
            technical_data['calibrationStandards'] = read_data('calibration_standards.json')
        except:
            technical_data['calibrationStandards'] = []

        try:
            technical_data['referrerMaster'] = read_data('referrer_master.json')
        except:
            technical_data['referrerMaster'] = []

        return jsonify(technical_data)
    except Exception as e:
        return jsonify({'message': f'Error fetching technical master data: {str(e)}'}), 500

# Referral Master Data CRUD Operations
@admin_bp.route('/api/admin/referral-master', methods=['GET'])
@token_required
def get_referral_master():
    """Get all referral master data"""
    try:
        # Try to read from referral pricing master file
        try:
            referral_data = read_data('referralPricingMaster.json')
            # Extract referral sources from the comprehensive data
            referral_sources = []
            if 'referralMaster' in referral_data:
                for ref_id, ref_data in referral_data['referralMaster'].items():
                    referral_sources.append(ref_data)

            return jsonify({
                'success': True,
                'data': referral_sources,
                'total': len(referral_sources)
            })
        except:
            # Return empty list if file doesn't exist
            return jsonify({
                'success': True,
                'data': [],
                'total': 0
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching referral master data: {str(e)}'
        }), 500

@admin_bp.route('/api/admin/referral-master', methods=['POST'])
@token_required
def add_referral_master():
    """Add new referral source"""
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        data = request.get_json()

        # Validate required fields - updated for new structure
        required_fields = ['id', 'name', 'description', 'referralType', 'email', 'phone', 'address']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({
                    'success': False,
                    'message': f'Missing required field: {field}'
                }), 400

        # Validate email format
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, data['email']):
            return jsonify({
                'success': False,
                'message': 'Invalid email format'
            }), 400

        # Validate referral type
        valid_types = ['Doctor', 'Hospital', 'Lab', 'Corporate', 'Insurance', 'Patient']
        if data['referralType'] not in valid_types:
            return jsonify({
                'success': False,
                'message': f'Invalid referral type. Must be one of: {", ".join(valid_types)}'
            }), 400

        # Validate type-specific fields
        type_specific_requirements = {
            'Doctor': ['specialization'],
            'Hospital': ['branch'],
            'Lab': ['accreditation'],
            'Corporate': ['registrationDetails'],
            'Insurance': ['policyCoverage'],
            'Patient': []  # patientReference is optional
        }

        referral_type = data['referralType']
        type_specific_fields = data.get('typeSpecificFields', {})

        for required_field in type_specific_requirements.get(referral_type, []):
            if required_field not in type_specific_fields or not type_specific_fields[required_field]:
                return jsonify({
                    'success': False,
                    'message': f'Missing required {referral_type}-specific field: {required_field}'
                }), 400

        # Load existing data
        try:
            referral_data = read_data('referralPricingMaster.json')
        except:
            # Initialize with default structure if file doesn't exist
            referral_data = {
                'referralMaster': {},
                'pricingSchemes': {},
                'testPricingMatrix': {},
                'discountRules': {},
                'commissionRules': {},
                'metadata': {
                    'version': '1.0',
                    'lastUpdated': datetime.now().isoformat(),
                    'updatedBy': request.current_user.get('id'),
                    'description': 'Comprehensive referral and pricing master configuration'
                }
            }

        # Check if referral ID already exists
        if data['id'] in referral_data['referralMaster']:
            return jsonify({
                'success': False,
                'message': f'Referral source with ID "{data["id"]}" already exists'
            }), 400

        # Auto-determine category based on referral type
        type_to_category = {
            'Doctor': 'medical',
            'Hospital': 'institutional',
            'Lab': 'institutional',
            'Corporate': 'corporate',
            'Insurance': 'insurance',
            'Patient': 'direct'
        }

        # Create new referral source with enhanced structure
        new_referral = {
            'id': data['id'],
            'name': data['name'],
            'description': data['description'],
            'referralType': data['referralType'],
            'category': type_to_category[data['referralType']],
            'defaultPricingScheme': data.get('defaultPricingScheme', 'standard'),
            'discountPercentage': float(data.get('discountPercentage', 0)),
            'commissionPercentage': float(data.get('commissionPercentage', 0)),
            'isActive': data.get('isActive', True),
            'priority': int(data.get('priority', 1)),
            # Common contact fields
            'email': data['email'],
            'phone': data['phone'],
            'address': data['address'],
            # Type-specific fields
            'typeSpecificFields': data.get('typeSpecificFields', {}),
            # Metadata
            'createdAt': datetime.now().isoformat(),
            'updatedAt': datetime.now().isoformat(),
            'createdBy': request.current_user.get('id')
        }

        # Add to referral master
        referral_data['referralMaster'][data['id']] = new_referral

        # Update metadata
        referral_data['metadata']['lastUpdated'] = datetime.now().isoformat()
        referral_data['metadata']['updatedBy'] = request.current_user.get('id')

        # Save to file
        write_data('referralPricingMaster.json', referral_data)

        return jsonify({
            'success': True,
            'message': 'Referral source added successfully',
            'data': new_referral
        }), 201

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error adding referral source: {str(e)}'
        }), 500

@admin_bp.route('/api/admin/referral-master/<referral_id>', methods=['PUT'])
@token_required
def update_referral_master(referral_id):
    """Update existing referral source"""
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        data = request.get_json()

        # Load existing data
        try:
            referral_data = read_data('referralPricingMaster.json')
        except:
            return jsonify({
                'success': False,
                'message': 'Referral master data file not found'
            }), 404

        # Check if referral exists
        if referral_id not in referral_data['referralMaster']:
            return jsonify({
                'success': False,
                'message': f'Referral source with ID "{referral_id}" not found'
            }), 404

        # Get existing referral
        existing_referral = referral_data['referralMaster'][referral_id]

        # Validate required fields for update
        if 'email' in data:
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, data['email']):
                return jsonify({
                    'success': False,
                    'message': 'Invalid email format'
                }), 400

        # Validate referral type if being updated
        if 'referralType' in data:
            valid_types = ['Doctor', 'Hospital', 'Lab', 'Corporate', 'Insurance', 'Patient']
            if data['referralType'] not in valid_types:
                return jsonify({
                    'success': False,
                    'message': f'Invalid referral type. Must be one of: {", ".join(valid_types)}'
                }), 400

        # Auto-determine category based on referral type
        type_to_category = {
            'Doctor': 'medical',
            'Hospital': 'institutional',
            'Lab': 'institutional',
            'Corporate': 'corporate',
            'Insurance': 'insurance',
            'Patient': 'direct'
        }

        # Update fields (preserve creation info)
        updatable_fields = [
            'name', 'description', 'referralType', 'defaultPricingScheme',
            'discountPercentage', 'commissionPercentage', 'isActive', 'priority',
            'email', 'phone', 'address', 'typeSpecificFields'
        ]

        for field in updatable_fields:
            if field in data:
                if field in ['discountPercentage', 'commissionPercentage']:
                    existing_referral[field] = float(data[field])
                elif field == 'priority':
                    existing_referral[field] = int(data[field])
                elif field == 'referralType':
                    existing_referral[field] = data[field]
                    # Auto-update category when referral type changes
                    existing_referral['category'] = type_to_category[data[field]]
                else:
                    existing_referral[field] = data[field]

        # Update metadata
        existing_referral['updatedAt'] = datetime.now().isoformat()
        referral_data['metadata']['lastUpdated'] = datetime.now().isoformat()
        referral_data['metadata']['updatedBy'] = request.current_user.get('id')

        # Save to file
        write_data('referralPricingMaster.json', referral_data)

        return jsonify({
            'success': True,
            'message': 'Referral source updated successfully',
            'data': existing_referral
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error updating referral source: {str(e)}'
        }), 500

@admin_bp.route('/api/admin/referral-master/<referral_id>', methods=['DELETE'])
@token_required
def delete_referral_master(referral_id):
    """Delete referral source"""
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        # Load existing data
        try:
            referral_data = read_data('referralPricingMaster.json')
        except:
            return jsonify({
                'success': False,
                'message': 'Referral master data file not found'
            }), 404

        # Check if referral exists
        if referral_id not in referral_data['referralMaster']:
            return jsonify({
                'success': False,
                'message': f'Referral source with ID "{referral_id}" not found'
            }), 404

        # Remove from referral master
        deleted_referral = referral_data['referralMaster'].pop(referral_id)

        # Update metadata
        referral_data['metadata']['lastUpdated'] = datetime.now().isoformat()
        referral_data['metadata']['updatedBy'] = request.current_user.get('id')

        # Save to file
        write_data('referralPricingMaster.json', referral_data)

        return jsonify({
            'success': True,
            'message': 'Referral source deleted successfully',
            'data': deleted_referral
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error deleting referral source: {str(e)}'
        }), 500

# Price Scheme Master Import for Lab-to-Lab Pricing
@admin_bp.route('/api/admin/price-scheme-master/import', methods=['POST'])
@token_required
def import_price_scheme_data():
    """Import lab-to-lab pricing data from Excel"""
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        data = request.get_json()
        import_data = data.get('data', [])

        if not import_data:
            return jsonify({
                'success': False,
                'message': 'No data provided for import'
            }), 400

        # Load existing price scheme data
        try:
            existing_data = read_data('price_scheme_master.json')
        except:
            existing_data = []

        # Get the next ID
        next_id = max([item.get('id', 0) for item in existing_data], default=0) + 1

        # Process and validate import data
        processed_data = []
        errors = []

        for index, row in enumerate(import_data):
            try:
                # Validate required fields
                required_fields = ['test_code', 'test_name', 'default_price', 'scheme_price']
                for field in required_fields:
                    if field not in row or not row[field]:
                        errors.append(f'Row {index + 1}: Missing required field {field}')
                        continue

                # Create processed record
                processed_record = {
                    'id': next_id,
                    'dept_code': row.get('dept_code', '@BC'),
                    'dept_name': row.get('dept_name', 'LAB'),
                    'scheme_code': row.get('scheme_code', '@000002'),
                    'scheme_name': row.get('scheme_name', 'L2L'),
                    'test_type': row.get('test_type', 'T'),
                    'test_code': str(row['test_code']),
                    'test_name': str(row['test_name']),
                    'default_price': float(row['default_price']),
                    'scheme_price': float(row['scheme_price']),
                    'price_percentage': float(row.get('price_percentage',
                        round((float(row['scheme_price']) / float(row['default_price'])) * 100, 2))),
                    'is_active': row.get('is_active', True),
                    'created_at': datetime.now().isoformat(),
                    'updated_at': datetime.now().isoformat(),
                    'created_by': request.current_user.get('id', 1)
                }

                processed_data.append(processed_record)
                next_id += 1

            except Exception as e:
                errors.append(f'Row {index + 1}: Error processing data - {str(e)}')

        if errors:
            return jsonify({
                'success': False,
                'message': 'Validation errors found',
                'errors': errors
            }), 400

        # Add processed data to existing data
        existing_data.extend(processed_data)

        # Save updated data
        write_data('price_scheme_master.json', existing_data)

        return jsonify({
            'success': True,
            'message': f'Successfully imported {len(processed_data)} records',
            'imported_count': len(processed_data),
            'total_records': len(existing_data)
        }), 201

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error importing price scheme data: {str(e)}'
        }), 500

@admin_bp.route('/api/admin/referrer-master-data', methods=['GET'])
@token_required
def get_referrer_master_data():
    """Get referrer master data (legacy endpoint for compatibility)"""
    try:
        # Try to read from JSON file first, then fallback to default data
        try:
            referrer_data = read_data('referrer_master.json')
        except:
            # Default referrer data structure
            referrer_data = {
                'referrerTypes': [
                    {'id': 'doctor', 'name': 'Doctor', 'description': 'Medical practitioners and physicians'},
                    {'id': 'self', 'name': 'Self', 'description': 'Self-referred patients'},
                    {'id': 'lab', 'name': 'Lab', 'description': 'Other laboratory facilities'},
                    {'id': 'hospital', 'name': 'Hospital', 'description': 'Hospital and medical institutions'},
                    {'id': 'corporate', 'name': 'Corporate', 'description': 'Corporate health programs'},
                    {'id': 'insurance', 'name': 'Insurance', 'description': 'Insurance companies and health plans'},
                    {'id': 'staff', 'name': 'Staff', 'description': 'Internal staff referrals'},
                    {'id': 'consultant', 'name': 'Consultant', 'description': 'Medical consultants and specialists'},
                    {'id': 'outstation', 'name': 'Outstation', 'description': 'Out-of-station referrals'}
                ],
                'referrerNames': {
                    'doctor': [
                        {'id': 'dr_001', 'name': 'Dr. Rajesh Kumar', 'specialization': 'Cardiologist'},
                        {'id': 'dr_002', 'name': 'Dr. Priya Sharma', 'specialization': 'Endocrinologist'},
                        {'id': 'dr_003', 'name': 'Dr. Amit Patel', 'specialization': 'General Physician'}
                    ],
                    'hospital': [
                        {'id': 'hosp_001', 'name': 'City General Hospital', 'type': 'Multi-specialty'},
                        {'id': 'hosp_002', 'name': 'Metro Medical Center', 'type': 'Super Specialty'}
                    ],
                    'lab': [
                        {'id': 'lab_001', 'name': 'PathLab Diagnostics', 'type': 'Full Service Lab'},
                        {'id': 'lab_002', 'name': 'QuickTest Laboratory', 'type': 'Express Testing'}
                    ],
                    'corporate': [
                        {'id': 'corp_001', 'name': 'TechCorp Health Program', 'company': 'TechCorp Industries'}
                    ],
                    'insurance': [
                        {'id': 'ins_001', 'name': 'HealthFirst Insurance', 'type': 'Health Insurance'}
                    ],
                    'staff': [
                        {'id': 'staff_001', 'name': 'Internal Medicine Department', 'department': 'Internal Medicine'}
                    ],
                    'consultant': [
                        {'id': 'cons_001', 'name': 'Dr. Expert Consultant', 'specialization': 'Radiology Consultant'}
                    ],
                    'outstation': [
                        {'id': 'out_001', 'name': 'Regional Medical Center', 'location': 'North Region'}
                    ],
                    'self': [
                        {'id': 'self_001', 'name': 'Walk-in Patient', 'description': 'Patient came directly without referral'}
                    ]
                }
            }

        return jsonify(referrer_data)
    except Exception as e:
        return jsonify({'message': f'Error fetching referrer master data: {str(e)}'}), 500

@admin_bp.route('/api/admin/technical-master-data/<category>', methods=['POST'])
@token_required
def add_technical_master_data_item(category):
    """Add new technical master data item"""
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    try:
        if category == 'resultMaster':
            return create_result_master_item(data)
        elif category == 'parameterMaster':
            return create_parameter_master_item(data)
        elif category == 'instrumentMaster':
            return create_instrument_master_item(data)
        elif category == 'reagentMaster':
            return create_reagent_master_item(data)
        else:
            return jsonify({'message': 'Invalid category'}), 400
    except Exception as e:
        return jsonify({'message': f'Error adding item: {str(e)}'}), 500

@admin_bp.route('/api/admin/technical-master-data/<category>/<int:item_id>', methods=['PUT'])
@token_required
def update_technical_master_data_item(category, item_id):
    """Update technical master data item"""
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    try:
        if category == 'resultMaster':
            return update_result_master_item(item_id, data)
        elif category == 'parameterMaster':
            return update_parameter_master_item(item_id, data)
        elif category == 'instrumentMaster':
            return update_instrument_master_item(item_id, data)
        elif category == 'reagentMaster':
            return update_reagent_master_item(item_id, data)
        else:
            return jsonify({'message': 'Invalid category'}), 400
    except Exception as e:
        return jsonify({'message': f'Error updating item: {str(e)}'}), 500

@admin_bp.route('/api/admin/technical-master-data/<category>/<int:item_id>', methods=['DELETE'])
@token_required
def delete_technical_master_data_item(category, item_id):
    """Delete technical master data item"""
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        if category == 'resultMaster':
            return delete_result_master_item(item_id)
        elif category == 'parameterMaster':
            return delete_parameter_master_item(item_id)
        elif category == 'instrumentMaster':
            return delete_instrument_master_item(item_id)
        elif category == 'reagentMaster':
            return delete_reagent_master_item(item_id)
        else:
            return jsonify({'message': 'Invalid category'}), 400
    except Exception as e:
        return jsonify({'message': f'Error deleting item: {str(e)}'}), 500

# Helper functions for technical master data
def create_result_master_item(data):
    """Create new result master item"""
    required_fields = ['result_name', 'parameter_name', 'test_name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    result_master = read_data('result_master.json')

    # Generate new ID
    new_id = 1
    if result_master:
        new_id = max(item['id'] for item in result_master) + 1

    # Create new result master item
    new_item = {
        'id': new_id,
        'result_name': data['result_name'],
        'parameter_name': data['parameter_name'],
        'test_name': data['test_name'],
        'unit': data.get('unit', ''),
        'result_type': data.get('result_type', 'numeric'),
        'reference_range': data.get('reference_range', ''),
        'normal_range': data.get('normal_range', ''),
        'critical_low': data.get('critical_low', ''),
        'critical_high': data.get('critical_high', ''),
        'decimal_places': data.get('decimal_places', 2),
        'calculation_formula': data.get('calculation_formula', ''),
        'validation_rules': data.get('validation_rules', ''),
        'display_order': data.get('display_order', 1),
        'is_calculated': data.get('is_calculated', False),
        'is_mandatory': data.get('is_mandatory', True),
        'allow_manual_entry': data.get('allow_manual_entry', True),
        'quality_control': data.get('quality_control', False),
        'instrument_id': data.get('instrument_id', ''),
        'method_id': data.get('method_id', ''),
        'specimen_type': data.get('specimen_type', ''),
        'reporting_unit': data.get('reporting_unit', ''),
        'conversion_factor': data.get('conversion_factor', 1.0),
        'interpretation_rules': data.get('interpretation_rules', ''),
        'panic_values': data.get('panic_values', ''),
        'delta_check_rules': data.get('delta_check_rules', ''),
        'age_specific_ranges': data.get('age_specific_ranges', ''),
        'gender_specific_ranges': data.get('gender_specific_ranges', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    result_master.append(new_item)
    write_data('result_master.json', result_master)

    return jsonify(new_item), 201

def create_parameter_master_item(data):
    """Create new parameter master item"""
    required_fields = ['parameter_name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    parameter_master = read_data('parameter_master.json')

    # Generate new ID
    new_id = 1
    if parameter_master:
        new_id = max(item['id'] for item in parameter_master) + 1

    # Create new parameter master item
    new_item = {
        'id': new_id,
        'parameter_name': data['parameter_name'],
        'code': data.get('code', ''),
        'unit': data.get('unit', ''),
        'parameter_type': data.get('parameter_type', 'Standard'),
        'category': data.get('category', 'General'),
        'description': data.get('description', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    parameter_master.append(new_item)
    write_data('parameter_master.json', parameter_master)

    return jsonify(new_item), 201

def create_instrument_master_item(data):
    """Create new instrument master item"""
    required_fields = ['instrument_name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    instrument_master = read_data('instrument_master.json')

    # Generate new ID
    new_id = 1
    if instrument_master:
        new_id = max(item['id'] for item in instrument_master) + 1

    # Create new instrument master item
    new_item = {
        'id': new_id,
        'instrument_name': data['instrument_name'],
        'model': data.get('model', ''),
        'manufacturer': data.get('manufacturer', ''),
        'serial_number': data.get('serial_number', ''),
        'department': data.get('department', ''),
        'description': data.get('description', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    instrument_master.append(new_item)
    write_data('instrument_master.json', instrument_master)

    return jsonify(new_item), 201

def create_reagent_master_item(data):
    """Create new reagent master item"""
    required_fields = ['reagent_name']
    for field in required_fields:
        if field not in data:
            return jsonify({'message': f'Missing required field: {field}'}), 400

    reagent_master = read_data('reagent_master.json')

    # Generate new ID
    new_id = 1
    if reagent_master:
        new_id = max(item['id'] for item in reagent_master) + 1

    # Create new reagent master item
    new_item = {
        'id': new_id,
        'reagent_name': data['reagent_name'],
        'code': data.get('code', ''),
        'lot_number': data.get('lot_number', ''),
        'expiry_date': data.get('expiry_date', ''),
        'manufacturer': data.get('manufacturer', ''),
        'is_active': data.get('is_active', True),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': request.current_user.get('id')
    }

    reagent_master.append(new_item)
    write_data('reagent_master.json', reagent_master)

    return jsonify(new_item), 201

# Update functions
def update_result_master_item(item_id, data):
    """Update result master item"""
    result_master = read_data('result_master.json')
    item_index = next((i for i, item in enumerate(result_master) if item['id'] == item_id), None)

    if item_index is None:
        return jsonify({'message': 'Result master item not found'}), 404

    item = result_master[item_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            item[key] = value

    item['updated_at'] = datetime.now().isoformat()
    write_data('result_master.json', result_master)
    return jsonify(item)

def update_parameter_master_item(item_id, data):
    """Update parameter master item"""
    parameter_master = read_data('parameter_master.json')
    item_index = next((i for i, item in enumerate(parameter_master) if item['id'] == item_id), None)

    if item_index is None:
        return jsonify({'message': 'Parameter master item not found'}), 404

    item = parameter_master[item_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            item[key] = value

    item['updated_at'] = datetime.now().isoformat()
    write_data('parameter_master.json', parameter_master)
    return jsonify(item)

def update_instrument_master_item(item_id, data):
    """Update instrument master item"""
    instrument_master = read_data('instrument_master.json')
    item_index = next((i for i, item in enumerate(instrument_master) if item['id'] == item_id), None)

    if item_index is None:
        return jsonify({'message': 'Instrument master item not found'}), 404

    item = instrument_master[item_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            item[key] = value

    item['updated_at'] = datetime.now().isoformat()
    write_data('instrument_master.json', instrument_master)
    return jsonify(item)

def update_reagent_master_item(item_id, data):
    """Update reagent master item"""
    reagent_master = read_data('reagent_master.json')
    item_index = next((i for i, item in enumerate(reagent_master) if item['id'] == item_id), None)

    if item_index is None:
        return jsonify({'message': 'Reagent master item not found'}), 404

    item = reagent_master[item_index]

    # Update only provided fields
    for key, value in data.items():
        if key not in ['id', 'created_at', 'created_by']:
            item[key] = value

    item['updated_at'] = datetime.now().isoformat()
    write_data('reagent_master.json', reagent_master)
    return jsonify(item)

# Delete functions
def delete_result_master_item(item_id):
    """Delete result master item"""
    result_master = read_data('result_master.json')
    item_index = next((i for i, item in enumerate(result_master) if item['id'] == item_id), None)

    if item_index is None:
        return jsonify({'message': 'Result master item not found'}), 404

    deleted_item = result_master.pop(item_index)
    write_data('result_master.json', result_master)
    return jsonify({'message': 'Result master item deleted successfully', 'deleted_item': deleted_item})

def delete_parameter_master_item(item_id):
    """Delete parameter master item"""
    parameter_master = read_data('parameter_master.json')
    item_index = next((i for i, item in enumerate(parameter_master) if item['id'] == item_id), None)

    if item_index is None:
        return jsonify({'message': 'Parameter master item not found'}), 404

    deleted_item = parameter_master.pop(item_index)
    write_data('parameter_master.json', parameter_master)
    return jsonify({'message': 'Parameter master item deleted successfully', 'deleted_item': deleted_item})

def delete_instrument_master_item(item_id):
    """Delete instrument master item"""
    instrument_master = read_data('instrument_master.json')
    item_index = next((i for i, item in enumerate(instrument_master) if item['id'] == item_id), None)

    if item_index is None:
        return jsonify({'message': 'Instrument master item not found'}), 404

    deleted_item = instrument_master.pop(item_index)
    write_data('instrument_master.json', instrument_master)
    return jsonify({'message': 'Instrument master item deleted successfully', 'deleted_item': deleted_item})

def delete_reagent_master_item(item_id):
    """Delete reagent master item"""
    reagent_master = read_data('reagent_master.json')
    item_index = next((i for i, item in enumerate(reagent_master) if item['id'] == item_id), None)

    if item_index is None:
        return jsonify({'message': 'Reagent master item not found'}), 404

    deleted_item = reagent_master.pop(item_index)
    write_data('reagent_master.json', reagent_master)
    return jsonify({'message': 'Reagent master item deleted successfully', 'deleted_item': deleted_item})

# Signature Management Routes
@admin_bp.route('/api/admin/signature', methods=['GET'])
@token_required
def get_current_signature():
    """Get current signature information"""
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        # Check if signature file exists
        signature_path = os.path.join('public', 'signature.jpeg')
        if os.path.exists(signature_path):
            # Get file stats
            stat = os.stat(signature_path)
            signature_info = {
                'url': '/signature.jpeg',
                'uploaded_at': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                'size': stat.st_size
            }
            return jsonify({
                'success': True,
                'data': signature_info
            })
        else:
            return jsonify({
                'success': True,
                'data': None,
                'message': 'No signature found'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error retrieving signature: {str(e)}'
        }), 500

@admin_bp.route('/api/admin/signature/upload', methods=['POST'])
@token_required
def upload_signature():
    """Upload a new signature image"""
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        # Check if file is present
        if 'signature' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No signature file provided'
            }), 400

        file = request.files['signature']
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': 'No file selected'
            }), 400

        # Validate file type
        allowed_extensions = {'jpg', 'jpeg', 'png'}
        file_extension = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''

        if file_extension not in allowed_extensions:
            return jsonify({
                'success': False,
                'message': 'Invalid file type. Please upload JPEG, JPG, or PNG files only.'
            }), 400

        # Validate file size (max 2MB)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)

        max_size = 2 * 1024 * 1024  # 2MB
        if file_size > max_size:
            return jsonify({
                'success': False,
                'message': 'File size exceeds 2MB limit'
            }), 400

        # Validate image using PIL if available
        if IMAGE_SUPPORT:
            try:
                img = PILImage.open(file)
                img.verify()
                file.seek(0)  # Reset file pointer after verification
            except Exception:
                return jsonify({
                    'success': False,
                    'message': 'Invalid image file'
                }), 400

        # Ensure public directory exists
        public_dir = 'public'
        if not os.path.exists(public_dir):
            os.makedirs(public_dir)

        # Save the file as signature.jpeg
        signature_path = os.path.join(public_dir, 'signature.jpeg')

        # If file is not JPEG, convert it
        if file_extension.lower() in ['png']:
            if IMAGE_SUPPORT:
                # Convert PNG to JPEG
                img = PILImage.open(file)
                if img.mode in ('RGBA', 'LA', 'P'):
                    # Convert to RGB for JPEG
                    background = PILImage.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'P':
                        img = img.convert('RGBA')
                    background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                    img = background
                img.save(signature_path, 'JPEG', quality=95)
            else:
                return jsonify({
                    'success': False,
                    'message': 'PNG conversion not supported. Please upload JPEG files.'
                }), 400
        else:
            # Save JPEG directly
            file.save(signature_path)

        return jsonify({
            'success': True,
            'message': 'Signature uploaded successfully',
            'data': {
                'url': '/signature.jpeg',
                'uploaded_at': datetime.now().isoformat()
            }
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error uploading signature: {str(e)}'
        }), 500

@admin_bp.route('/api/admin/signature', methods=['DELETE'])
@token_required
def remove_signature():
    """Remove the current signature"""
    # Check if user has admin privileges
    if request.current_user.get('role') not in ['admin', 'hub_admin']:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        signature_path = os.path.join('public', 'signature.jpeg')

        if os.path.exists(signature_path):
            os.remove(signature_path)
            return jsonify({
                'success': True,
                'message': 'Signature removed successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'No signature found to remove'
            }), 404

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error removing signature: {str(e)}'
        }), 500

# Excel Data Integration Routes
@admin_bp.route('/api/admin/excel-data', methods=['GET'])
@token_required
def get_excel_data():
    """Get all Excel imported data"""
    try:
        data = read_data('excel_test_data.json')
        return jsonify({'data': data})
    except Exception as e:
        return jsonify({'message': f'Failed to fetch Excel data: {str(e)}'}), 500

@admin_bp.route('/api/admin/excel-data/search', methods=['GET'])
@token_required
def search_excel_data():
    """Search Excel data by test name or code"""
    try:
        query = request.args.get('q', '').strip().lower()
        if not query:
            return jsonify({'data': []})

        data = read_data('excel_test_data.json')

        # Search by test name or test code
        results = []
        for item in data:
            test_name = str(item.get('test_name', '')).lower()
            test_code = str(item.get('test_code', '')).lower()

            if query in test_name or query in test_code:
                results.append(item)

        return jsonify({'data': results})
    except Exception as e:
        return jsonify({'message': f'Search failed: {str(e)}'}), 500

@admin_bp.route('/api/admin/excel-data/lookup/<test_code>', methods=['GET'])
@token_required
def lookup_test_by_code(test_code):
    """Lookup test data by test code for auto-population"""
    try:
        # Format test code to 6 digits
        formatted_code = f"{int(test_code):06d}" if test_code.isdigit() else test_code

        data = read_data('excel_test_data.json')

        # Find test by code
        for item in data:
            if item.get('test_code') == formatted_code:
                return jsonify({'data': item, 'found': True})

        return jsonify({'data': None, 'found': False})
    except Exception as e:
        return jsonify({'message': f'Lookup failed: {str(e)}'}), 500

@admin_bp.route('/api/admin/excel-data/lookup-by-name/<test_name>', methods=['GET'])
@token_required
def lookup_test_by_name(test_name):
    """Lookup test data by test name for auto-population"""
    try:
        data = read_data('excel_test_data.json')

        # Find test by name (case insensitive)
        test_name_lower = test_name.lower()
        for item in data:
            if item.get('test_name', '').lower() == test_name_lower:
                return jsonify({'data': item, 'found': True})

        return jsonify({'data': None, 'found': False})
    except Exception as e:
        return jsonify({'message': f'Lookup failed: {str(e)}'}), 500

# Enhanced Test Master Routes
@admin_bp.route('/api/admin/test-master-enhanced', methods=['GET'])
@token_required
def get_enhanced_test_master():
    """Get enhanced test master data"""
    try:
        data = read_data('test_master_enhanced.json')
        return jsonify({'data': data})
    except Exception as e:
        return jsonify({'message': f'Failed to fetch enhanced test master: {str(e)}'}), 500

@admin_bp.route('/api/admin/test-master-enhanced', methods=['POST'])
@token_required
def add_enhanced_test_master():
    """Add new enhanced test master entry"""
    try:
        data = request.get_json()

        # Read existing data
        existing_data = read_data('test_master_enhanced.json')

        # Generate new ID
        new_id = max([item.get('id', 0) for item in existing_data], default=0) + 1

        # Add metadata
        data['id'] = new_id
        data['created_at'] = datetime.now().isoformat()
        data['updated_at'] = datetime.now().isoformat()
        data['excel_source'] = False  # Manual entry

        # Add to existing data
        existing_data.append(data)

        # Save
        write_data('test_master_enhanced.json', existing_data)

        return jsonify({'data': data, 'message': 'Test master entry added successfully'})
    except Exception as e:
        return jsonify({'message': f'Failed to add test master entry: {str(e)}'}), 500

# Enhanced Result Master Routes
@admin_bp.route('/api/admin/result-master-enhanced', methods=['GET'])
@token_required
def get_enhanced_result_master():
    """Get enhanced result master data"""
    try:
        data = read_data('result_master_enhanced.json')
        return jsonify({'data': data})
    except Exception as e:
        return jsonify({'message': f'Failed to fetch enhanced result master: {str(e)}'}), 500

@admin_bp.route('/api/admin/result-master-enhanced', methods=['POST'])
@token_required
def add_enhanced_result_master():
    """Add new enhanced result master entry"""
    try:
        data = request.get_json()

        # Read existing data
        existing_data = read_data('result_master_enhanced.json')

        # Generate new ID
        new_id = max([item.get('id', 0) for item in existing_data], default=0) + 1

        # Add metadata
        data['id'] = new_id
        data['created_at'] = datetime.now().isoformat()
        data['updated_at'] = datetime.now().isoformat()
        data['excel_source'] = False  # Manual entry

        # Add to existing data
        existing_data.append(data)

        # Save
        write_data('result_master_enhanced.json', existing_data)

        return jsonify({'data': data, 'message': 'Result master entry added successfully'})
    except Exception as e:
        return jsonify({'message': f'Failed to add result master entry: {str(e)}'}), 500
    
    
    
@admin_bp.route('/api/admin/test-master-enhanced/<int:item_id>', methods=['PUT'])
@token_required
def update_enhanced_test_master(item_id):
    """Update an existing test master entry by ID"""
    try:
        updated_data = request.get_json()
        existing_data = read_data('test_master_enhanced.json')

        # Find and update the matching item
        for index, item in enumerate(existing_data):
            if item.get('id') == item_id:
                updated_data['id'] = item_id  # Preserve original ID
                updated_data['created_at'] = item.get('created_at')  # Preserve original creation time
                updated_data['updated_at'] = datetime.now().isoformat()
                updated_data['excel_source'] = item.get('excel_source', False)

                # Update the item
                existing_data[index] = updated_data

                # Save the updated data
                write_data('test_master_enhanced.json', existing_data)

                return jsonify({
                    'data': updated_data,
                    'message': 'Test master entry updated successfully'
                }), 200

        return jsonify({'message': 'Test master entry not found'}), 404

    except Exception as e:
        return jsonify({'message': f'Failed to update test master entry: {str(e)}'}), 500


@admin_bp.route('/api/admin/result-master-enhanced/<int:item_id>', methods=['PUT'])
@token_required
def update_enhanced_result_master(item_id):
    """Update an existing result master entry by ID"""
    try:
        updated_data = request.get_json()
        existing_data = read_data('result_master_enhanced.json')

        # Find item by ID
        for index, item in enumerate(existing_data):
            if item.get('id') == item_id:
                updated_data['id'] = item_id  # Ensure ID stays unchanged
                updated_data['created_at'] = item.get('created_at')  # Preserve original
                updated_data['updated_at'] = datetime.now().isoformat()
                existing_data[index] = updated_data

                # Save updated data
                write_data('result_master_enhanced.json', existing_data)

                return jsonify({
                    'data': updated_data,
                    'message': 'Result master entry updated successfully'
                }), 200

        return jsonify({'message': 'Result master entry not found'}), 404

    except Exception as e:
        return jsonify({'message': f'Failed to update result master entry: {str(e)}'}), 500

# Price Scheme Master Routes
@admin_bp.route('/api/admin/price-scheme-master', methods=['GET'])
@token_required
def get_price_scheme_master():
    """Get all price scheme master data"""
    try:
        # Check if user has admin privileges
        if request.current_user.get('role') not in ['admin', 'hub_admin']:
            return jsonify({'message': 'Unauthorized'}), 403

        data = read_data('price_scheme_master.json')
        return jsonify({'data': data})
    except Exception as e:
        return jsonify({'message': f'Failed to fetch price scheme master: {str(e)}'}), 500

@admin_bp.route('/api/admin/price-scheme-master', methods=['POST'])
@token_required
def add_price_scheme_master():
    """Add new price scheme master entry"""
    try:
        # Check if user has admin privileges
        if request.current_user.get('role') not in ['admin', 'hub_admin']:
            return jsonify({'message': 'Unauthorized'}), 403

        data = request.get_json()

        # Validate required fields
        required_fields = ['test_code', 'test_name', 'scheme_code', 'default_price']
        for field in required_fields:
            if field not in data or data[field] in [None, ""]:
                return jsonify({'message': f'Missing required field: {field}'}), 400

        # Read existing data
        existing_data = read_data('price_scheme_master.json')

        # Generate new ID
        new_id = max([item.get('id', 0) for item in existing_data], default=0) + 1

        # Calculate price percentage if scheme_price is provided
        default_price = float(data['default_price'])
        scheme_price = float(data.get('scheme_price', default_price))
        price_percentage = round((scheme_price / default_price) * 100, 2) if default_price > 0 else 0

        # Add metadata
        new_entry = {
            'id': new_id,
            'dept_code': data.get('dept_code', ''),
            'dept_name': data.get('dept_name', ''),
            'scheme_code': data['scheme_code'],
            'scheme_name': data.get('scheme_name', ''),
            'test_type': data.get('test_type', 'T'),
            'test_code': data['test_code'],
            'test_name': data['test_name'],
            'default_price': default_price,
            'scheme_price': scheme_price,
            'price_percentage': price_percentage,
            'is_active': data.get('is_active', True),
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'created_by': request.current_user.get('id', 1)
        }

        # Add to existing data
        existing_data.append(new_entry)

        # Save
        write_data('price_scheme_master.json', existing_data)

        return jsonify({'data': new_entry, 'message': 'Price scheme entry added successfully'})
    except Exception as e:
        return jsonify({'message': f'Failed to add price scheme entry: {str(e)}'}), 500

@admin_bp.route('/api/admin/price-scheme-master/<int:item_id>', methods=['PUT'])
@token_required
def update_price_scheme_master(item_id):
    """Update an existing price scheme master entry by ID"""
    try:
        # Check if user has admin privileges
        if request.current_user.get('role') not in ['admin', 'hub_admin']:
            return jsonify({'message': 'Unauthorized'}), 403

        updated_data = request.get_json()
        existing_data = read_data('price_scheme_master.json')

        # Find the item to update
        item_index = next((i for i, item in enumerate(existing_data) if item.get('id') == item_id), None)
        if item_index is None:
            return jsonify({'message': 'Price scheme entry not found'}), 404

        # Update the item
        item = existing_data[item_index]

        # Update fields
        for key, value in updated_data.items():
            if key not in ['id', 'created_at', 'created_by']:
                item[key] = value

        # Recalculate price percentage if prices changed
        if 'default_price' in updated_data or 'scheme_price' in updated_data:
            default_price = float(item.get('default_price', 0))
            scheme_price = float(item.get('scheme_price', default_price))
            item['price_percentage'] = round((scheme_price / default_price) * 100, 2) if default_price > 0 else 0

        item['updated_at'] = datetime.now().isoformat()

        # Save
        write_data('price_scheme_master.json', existing_data)

        return jsonify({'data': item, 'message': 'Price scheme entry updated successfully'})
    except Exception as e:
        return jsonify({'message': f'Failed to update price scheme entry: {str(e)}'}), 500

@admin_bp.route('/api/admin/price-scheme-master/<int:item_id>', methods=['DELETE'])
@token_required
def delete_price_scheme_master(item_id):
    """Delete a price scheme master entry by ID"""
    try:
        # Check if user has admin privileges
        if request.current_user.get('role') not in ['admin', 'hub_admin']:
            return jsonify({'message': 'Unauthorized'}), 403

        existing_data = read_data('price_scheme_master.json')

        # Find the item to delete
        item_index = next((i for i, item in enumerate(existing_data) if item.get('id') == item_id), None)
        if item_index is None:
            return jsonify({'message': 'Price scheme entry not found'}), 404

        # Remove the item
        deleted_item = existing_data.pop(item_index)

        # Save
        write_data('price_scheme_master.json', existing_data)

        return jsonify({'message': 'Price scheme entry deleted successfully'})
    except Exception as e:
        return jsonify({'message': f'Failed to delete price scheme entry: {str(e)}'}), 500

# Schemes Master Routes (for dropdown data)
@admin_bp.route('/api/admin/schemes-master', methods=['GET'])
@token_required
def get_schemes_master():
    """Get all schemes master data for dropdowns"""
    try:
        data = read_data('schemes_master.json')
        return jsonify({'data': data})
    except Exception as e:
        return jsonify({'message': f'Failed to fetch schemes master: {str(e)}'}), 500

@admin_bp.route('/api/admin/schemes-master', methods=['POST'])
@token_required
def add_schemes_master():
    """Add new scheme master entry"""
    try:
        # Check if user has admin privileges
        if request.current_user.get('role') not in ['admin', 'hub_admin']:
            return jsonify({'message': 'Unauthorized'}), 403

        data = request.get_json()

        # Validate required fields
        required_fields = ['scheme_code', 'scheme_name']
        for field in required_fields:
            if field not in data or data[field] in [None, ""]:
                return jsonify({'message': f'Missing required field: {field}'}), 400

        # Read existing data
        existing_data = read_data('schemes_master.json')

        # Check for duplicate scheme code
        if any(item.get('scheme_code') == data['scheme_code'] for item in existing_data):
            return jsonify({'message': 'Scheme code already exists'}), 400

        # Generate new ID
        new_id = max([item.get('id', 0) for item in existing_data], default=0) + 1

        # Add metadata
        new_entry = {
            'id': new_id,
            'scheme_code': data['scheme_code'],
            'scheme_name': data['scheme_name'],
            'description': data.get('description', ''),
            'is_active': data.get('is_active', True),
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat(),
            'created_by': request.current_user.get('id', 1)
        }

        # Add to existing data
        existing_data.append(new_entry)

        # Save
        write_data('schemes_master.json', existing_data)

        return jsonify({'data': new_entry, 'message': 'Scheme added successfully'})
    except Exception as e:
        return jsonify({'message': f'Failed to add scheme: {str(e)}'}), 500
